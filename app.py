
# this is okay hooo
from flask import Flask, request, jsonify, render_template, send_file, Response, redirect, url_for, send_from_directory
import socket
from zeroconf import ServiceInfo, Zeroconf
from datetime import datetime, timedelta
import calendar
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from collections import deque
import time
import json
import sys
import uuid
import re
import shutil
import glob
from werkzeug.utils import secure_filename
import secrets
import hashlib

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

app = Flask(__name__)

# Disable caching for development
@app.after_request
def add_no_cache_headers(response):
    response.cache_control.no_cache = True
    response.cache_control.no_store = True
    response.cache_control.must_revalidate = True
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# Note: resource files (audio, gifs, etc.) are served by a later,
# more robust `serve_resource_file` implementation which performs
# additional safety checks and logging.

# File paths
STUDENTS_JSON = "students.json"
ATTENDANCE_LOG_CSV = "attendance_log.csv"
PHOTOS_DIR = "photos"
# Persistent counters to avoid reusing sequential IDs when students are deleted
ID_COUNTERS_FILE = "id_counters.json"
ALLOWED_PHOTO_EXTENSIONS = {'jpg', 'jpeg', 'png', 'webp', 'avif'}
MAX_PHOTO_SIZE = 5 * 1024 * 1024  # 5MB

# Default profile images (stored in the `resource/` folder)
DEFAULT_PHOTOS = {
    'male': 'boyDefaultProfile.jpg',
    'female': 'girlDefaultProfile.jpg',
    'monk': 'monkDefaltProfile.jpg'
}

# Ensure photos directory exists under the project directory
PROJECT_ROOT = os.path.dirname(__file__)
PHOTOS_DIR_PATH = os.path.join(PROJECT_ROOT, PHOTOS_DIR)
if not os.path.exists(PHOTOS_DIR_PATH):
    os.makedirs(PHOTOS_DIR_PATH, exist_ok=True)

# Copy default avatars into photos/ if they don't already exist there
for key, filename in DEFAULT_PHOTOS.items():
    src = os.path.join(PROJECT_ROOT, 'resource', filename)
    dst = os.path.join(PHOTOS_DIR_PATH, filename)
    if os.path.exists(src) and not os.path.exists(dst):
        try:
            shutil.copyfile(src, dst)
        except Exception as e:
            print(f"⚠️ Could not copy default photo {filename}: {e}")

# Memory for Kiosk and Registration Capture
LAST_SCAN_DATA = {"uid": None, "name": None, "status": None, "timestamp": None, "message": None, "sound": None}
PENDING_REG_DATA = None  # Holds form data while waiting for a physical card tap

# Registration Confirmation Tracking
PENDING_CONFIRMATIONS = {}  # Dict: {session_id: {"card_uid": uid, "student_id": id, "name": name, "confirmed": False}}
CURRENT_REGISTRATION_SESSION_ID = None  # ID to track current registration attempt

# Duplicate Error Tracking - to prevent stale errors from showing repeatedly
DUPLICATE_ERROR_TIMESTAMP = None  # Timestamp when duplicate error was last set

# Real-time Stream Storage - Keep last 50 check-ins
RECENT_CHECKINS = deque(maxlen=50)
LAST_REGISTRATION = None  # Most recent registration event
SSE_CLIENTS = []  # List of SSE client generators

# PN532 Status Tracking
PN532_STATUS = {
    "status": "disconnected",  # connected, disconnected, reconnecting
    "last_heartbeat": None,
    "esp32_ip": None
}

# Class Schedule Configuration
# Define how many sessions each class has and their time windows
CLASS_CONFIG = {
    "Teacher": {
        "sessions": 2,
        "name": "គ្រូបង្រៀន (2 sessions)",
        "schedule_times": {
            "session_1": {"checkin_start": "06:50", "checkin_end": "08:30", "checkout_start": "07:30", "checkout_end": "09:20"},
            "session_2": {"checkin_start": "13:50", "checkin_end": "15:30", "checkout_start": "14:30", "checkout_end": "17:20"}
        }
    },
    "Grade 1": {
        "sessions": 2,
        "name": "ថ្នាក់ទី១ (2 sessions)",
        "schedule_times": {
            "session_1": {"checkin_start": "06:50", "checkin_end": "08:30", "checkout_start": "07:30", "checkout_end": "09:20"},
            "session_2": {"checkin_start": "13:50", "checkin_end": "15:30", "checkout_start": "14:30", "checkout_end": "17:20"}
        }
    },
    "Grade 2": {
        "sessions": 2,
        "name": "ថ្នាក់ទី២ (2 sessions)",
        "schedule_times": {
            "session_1": {"checkin_start": "06:50", "checkin_end": "08:30", "checkout_start": "07:30", "checkout_end": "09:20"},
            "session_2": {"checkin_start": "13:50", "checkin_end": "15:30", "checkout_start": "14:30", "checkout_end": "17:20"}
        }
    },
    "Grade 3": {
        "sessions": 2,
        "name": "ថ្នាក់ទី៣ (2 sessions)",
        "schedule_times": {
            "session_1": {"checkin_start": "06:50", "checkin_end": "08:30", "checkout_start": "07:30", "checkout_end": "09:20"},
            "session_2": {"checkin_start": "13:50", "checkin_end": "15:30", "checkout_start": "14:30", "checkout_end": "17:20"}
        }
    },
    "Kindergarten": {
        "sessions": 2,
        "name": "ថ្នាក់មត្តេយ្យ (2 sessions)",
        "schedule_times": {
            "session_1": {"checkin_start": "06:50", "checkin_end": "08:30", "checkout_start": "07:30", "checkout_end": "09:20"},
            "session_2": {"checkin_start": "13:50", "checkin_end": "15:30", "checkout_start": "14:30", "checkout_end": "17:20"}
        }
    },
    "Computer Class - Level 1": {
        "sessions": 1,
        "name": "ថ្នាក់កុំព្យូទ័រ ម៉ោង​៥-៦ ល្ងាច ចន្ទ-សុក្រ",
        "schedule_times": {
            "session_1": {"checkin_start": "16:50", "checkin_end": "17:27", "checkout_start": "17:32", "checkout_end": "18:20"}
        }
    },
    "Computer Class - Level 2": {
        "sessions": 1,
        "name": "ថ្នាក់កុំព្យូទ័រ ម៉ោង​៧-៨ ល្ងាច​ ចន្ទ-សុក្រ",
        "schedule_times": {
            "session_1": {"checkin_start": "17:50", "checkin_end": "18:27", "checkout_start": "18:32", "checkout_end": "19:20"}
      }
    },
    "Computer Class - Level A1": {
        "sessions": 1,
        "name": "ថ្នាក់កុំព្យូទ័រ ម៉ោង​៥-៦ ល្ងាច សៅរ៍-អាទិត្យ",
        "schedule_times": {
            "session_1": {"checkin_start": "16:50", "checkin_end": "17:27", "checkout_start": "17:32", "checkout_end": "18:20"}
        }
    },
    "Computer Class - Level A2": {
        "sessions": 1,
        "name": "ថ្នាក់កុំព្យូទ័រ ម៉ោង​៧-៨ ល្ងាច សៅរ៍-អាទិត្យ",
        "schedule_times": {
            "session_1": {"checkin_start": "17:50", "checkin_end": "18:27", "checkout_start": "18:32", "checkout_end": "19:20"}
        }
    },
    "17:00-18:00​ ចន្ទ​ សុក្រ": {
        "sessions": 1,
        "name": "ថ្នាក់កុំព្យូទ័រ ម៉ោង​៥-៦ ល្ងាច ចន្ទ-សុក្រ",
        "schedule_times": {
            "session_1": {"checkin_start": "16:50", "checkin_end": "17:27", "checkout_start": "17:32", "checkout_end": "18:20"}
        }
    },
    "18:00-19:00​ ចន្ទ​ សុក្រ": {
        "sessions": 1,
        "name": "ថ្នាក់កុំព្យូទ័រ ម៉ោង​៧-៨ ល្ងាច​ ចន្ទ-សុក្រ",
        "schedule_times": {
            "session_1": {"checkin_start": "17:50", "checkin_end": "18:27", "checkout_start": "18:32", "checkout_end": "19:20"}
        }
    },
    "17:00-18:00​ សៅរ៍ អាទិត្យ": {
        "sessions": 1,
        "name": "ថ្នាក់កុំព្យូទ័រ ម៉ោង​៥-៦ ល្ងាច សៅរ៍-អាទិត្យ",
        "schedule_times": {
            "session_1": {"checkin_start": "16:50", "checkin_end": "17:27", "checkout_start": "17:32", "checkout_end": "18:20"}
        }
    },
    "18:00-19:00​ សៅរ៍ អាទិត្យ": {
        "sessions": 1,
        "name": "ថ្នាក់កុំព្យូទ័រ ម៉ោង​៧-៨ ល្ងាច សៅរ៍-អាទិត្យ",
        "schedule_times": {
            "session_1": {"checkin_start": "17:50", "checkin_end": "18:27", "checkout_start": "18:32", "checkout_end": "19:20"}
        }
    },
    "English Class - Essential": {
        "sessions": 1,
        "name": "ថ្នាក់ភាសាអង់គ្លេស - Essential",
        "schedule_times": {
            "session_1": {"checkin_start": "17:50", "checkin_end": "18:27", "checkout_start": "18:32", "checkout_end": "19:20"}
        }
    },
    "English Class - Beginner": {
        "sessions": 1,
        "name": "ថ្នាក់ភាសាអង់គ្លេស - Beginner",
        "schedule_times": {
            "session_1": {"checkin_start": "18:50", "checkin_end": "19:27", "checkout_start": "19:32", "checkout_end": "20:20"}
        }
    }
}

# ID Prefix Mapping: Each class/level combo gets a unique numeric prefix
# This ensures consistent ID generation across all class types
CLASS_ID_MAP = {
    "Teacher": 0,
    "Grade 1": 1,
    "Grade 2": 2,
    "Grade 3": 3,
    "Kindergarten": 4,
    "Computer Class - Level 1": 5,      # 17:00-18:00 weekday
    "Computer Class - Level 2": 6,      # 18:00-19:00 weekday
    "Computer Class - Level A1": 7,     # 17:00-18:00 weekend
    "Computer Class - Level A2": 8,     # 18:00-19:00 weekend
    "English Class - Essential": 9,
    "English Class - Beginner": 10
}

# Localized schedule keys mapping to same ID prefixes
CLASS_ID_MAP.update({
    "17:00-18:00​ ចន្ទ​ សុក្រ": CLASS_ID_MAP.get("Computer Class - Level 1", 5),
    "18:00-19:00​ ចន្ទ​ សុក្រ": CLASS_ID_MAP.get("Computer Class - Level 2", 6),
    "17:00-18:00​ សៅរ៍ អាទិត្យ": CLASS_ID_MAP.get("Computer Class - Level A1", 7),
    "18:00-19:00​ សៅរ៍ អាទិត្យ": CLASS_ID_MAP.get("Computer Class - Level A2", 8)
})

def schedule_is_computer_class(schedule: str) -> bool:
    """Return True if the given schedule string refers to a Computer Class.
    Accepts either English class names that contain 'Computer Class' or
    localized schedule strings like '17:00-18:00​ សៅរ៍ អាទិត្យ' / '18:00-19:00​ ចន្ទ​ សុក្រ'.
    """
    if not schedule:
        return False
    if isinstance(schedule, str) and "Computer Class" in schedule:
        return True
    # Detect time-based computer class schedules (17:00-18:00 or 18:00-19:00)
    if any(t in schedule for t in ("17:00-18:00", "18:00-19:00")):
        # also ensure it contains a Khmer weekday token common to class labels
        khmer_days = ["សៅរ៍", "អាទិត្យ", "ចន្ទ", "សុក្រ"]
        if any(d in schedule for d in khmer_days):
            return True
    return False

# Minimum delay (in minutes) between check-in and check-out
MIN_CHECKOUT_DELAY_MINUTES = 30

# Grace window (minutes) to allow late check-ins after checkin_end
GRACE_WINDOW_MINUTES = int(os.environ.get('GRACE_WINDOW_MINUTES', '30'))
# Admin keys: comma-separated tokens or token:username entries (env var)
ADMIN_KEYS_RAW = os.environ.get('ADMIN_KEYS', 'admin')
ADMIN_KEYS = {}
for item in ADMIN_KEYS_RAW.split(','):
    item = item.strip()
    if not item:
        continue
    if ':' in item:
        token, name = item.split(':', 1)
        ADMIN_KEYS[token.strip()] = name.strip()
    else:
        ADMIN_KEYS[item] = item[:8]

# Audit log for manual edits
AUDIT_LOG_CSV = 'attendance_audit.csv'

# Persistent admin keys file
ADMIN_KEYS_FILE = 'admin_keys.json'

def load_persistent_admin_keys():
    if not os.path.exists(ADMIN_KEYS_FILE):
        return []
    try:
        with open(ADMIN_KEYS_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception as e:
        print(f"⚠️ Could not load admin keys file: {e}")
        return []

def save_persistent_admin_keys(keys):
    try:
        with open(ADMIN_KEYS_FILE, 'w', encoding='utf-8') as f:
            json.dump(keys, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ Could not save admin keys file: {e}")

def ensure_admin_keys_file():
    """Migrate in-memory ENV keys into persistent file if file missing."""
    if os.path.exists(ADMIN_KEYS_FILE):
        return
    entries = []
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for token, name in ADMIN_KEYS.items():
        token_hash = hashlib.sha256(token.encode()).hexdigest()
        entries.append({'id': str(uuid.uuid4()), 'username': name, 'token_hash': token_hash, 'created_at': now})
    save_persistent_admin_keys(entries)

ensure_admin_keys_file()

def extract_admin_token():
    """Extract admin token from Authorization header (Bearer/Basic) or query/json param."""
    auth = request.headers.get('Authorization')
    if auth:
        if auth.startswith('Bearer '):
            return auth.split(None, 1)[1].strip()
        if auth.startswith('Basic '):
            try:
                import base64
                payload = base64.b64decode(auth.split()[1]).decode('utf-8')
                # payload is 'user:password' — treat password as token
                parts = payload.split(':', 1)
                if len(parts) == 2:
                    return parts[1]
            except Exception:
                return None

    # Query param or JSON body
    token = request.args.get('admin_token') or request.args.get('token')
    if not token and request.is_json:
        token = (request.get_json(silent=True) or {}).get('admin_token')
    return token

def get_admin_user(token):
    """Return admin username for a given token or None."""
    if not token:
        return None
    # Check ENV keys first
    if token in ADMIN_KEYS:
        return ADMIN_KEYS.get(token)

    # Check persistent keys (hash compare)
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    keys = load_persistent_admin_keys()
    for entry in keys:
        if hmac_compare(entry.get('token_hash', ''), token_hash):
            return entry.get('username')
    return None

def hmac_compare(a, b):
    try:
        # Use hmac.compare_digest if available
        import hmac as _h
        return _h.compare_digest(a, b)
    except Exception:
        return a == b

def is_admin(token):
    """Deprecated: keep for compatibility. Use get_admin_user instead."""
    return bool(get_admin_user(token))

# Admin rate limiting (requests per minute per token)
ADMIN_RATE_LIMIT_PER_MINUTE = int(os.environ.get('ADMIN_RATE_LIMIT_PER_MINUTE', '60'))
ADMIN_REQUEST_LOG = {}

def is_admin(token):
    """Check admin token (simple secret-based auth)."""
    return bool(token) and token == ADMIN_SECRET


def check_admin_rate(token):
    """Simple per-token rate limiter. Returns (allowed:bool, remaining:int)."""
    now_ts = time.time()
    window = 60
    max_req = ADMIN_RATE_LIMIT_PER_MINUTE
    lst = ADMIN_REQUEST_LOG.get(token, [])
    # keep only recent timestamps
    lst = [t for t in lst if now_ts - t < window]
    if len(lst) >= max_req:
        ADMIN_REQUEST_LOG[token] = lst
        return False, 0
    lst.append(now_ts)
    ADMIN_REQUEST_LOG[token] = lst
    return True, max_req - len(lst)

def append_audit_entry(admin, student_id, date_str, slot, old_value, new_value, reason=''):
    """Append a single audit row to AUDIT_LOG_CSV"""
    header = ['Timestamp', 'Admin', 'StudentID', 'Date', 'Slot', 'OldValue', 'NewValue', 'Reason']
    row = [datetime.now().strftime('%Y-%m-%d %H:%M:%S'), admin, student_id, date_str, slot, old_value, new_value, reason]
    try:
        file_exists = os.path.exists(AUDIT_LOG_CSV)
        with open(AUDIT_LOG_CSV, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(header)
            writer.writerow(row)
    except Exception as e:
        print(f"⚠️ Audit log write failed: {e}")


def init_student_data():
    """Initialize or migrate student data to JSON format"""
    if not os.path.exists(STUDENTS_JSON):
        # Check if old students.csv exists and migrate it
        if os.path.exists("students.csv"):
            print("📝 Migrating students.csv to students.json...")
            try:
                students = []
                seen_student_ids = set()
                
                with open("students.csv", 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        student_id = row.get('StudentID', row.get('UID', ''))
                        if not student_id or student_id in seen_student_ids:
                            continue  # Skip duplicates or empty
                        
                        seen_student_ids.add(student_id)
                        card_uid = row.get('CardUID', row.get('UID', ''))
                        
                        student = {
                            "student_id": student_id,
                            "name": row.get('Name', ''),
                            "sex": row.get('Sex', ''),
                            "schedule": row.get('Schedule', ''),
                            "card_uids": [card_uid] if card_uid else []
                        }
                        students.append(student)
                
                with open(STUDENTS_JSON, 'w', encoding='utf-8') as f:
                    json.dump(students, f, ensure_ascii=False, indent=2)
                print(f"✅ Migration complete! {len(students)} students migrated to JSON.")
            except Exception as e:
                print(f"⚠️ Migration failed: {e}. Creating empty students.json")
                with open(STUDENTS_JSON, 'w', encoding='utf-8') as f:
                    json.dump([], f, ensure_ascii=False, indent=2)
        else:
            # Create empty students.json if it doesn't exist
            with open(STUDENTS_JSON, 'w', encoding='utf-8') as f:
                json.dump([], f, ensure_ascii=False, indent=2)
    
    if not os.path.exists(ATTENDANCE_LOG_CSV):
        with open(ATTENDANCE_LOG_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # Changed UID to StudentID for tracking by student instead of card
            writer.writerow(['Date', 'StudentID', 'Name', 'In1', 'Out1', 'In2', 'Out2', 'Status'])

def load_students():
    """Load all students from JSON"""
    try:
        with open(STUDENTS_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"⚠️ Error loading students: {e}")
        return []

def save_students(students):
    """Save students to JSON"""
    try:
        with open(STUDENTS_JSON, 'w', encoding='utf-8') as f:
            json.dump(students, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"❌ Error saving students: {e}")


def load_id_counters():
    """Load per-class next-id counters from disk."""
    if not os.path.exists(ID_COUNTERS_FILE):
        return {}
    try:
        with open(ID_COUNTERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"⚠️ Could not load ID counters: {e}")
        return {}


def save_id_counters(counters):
    """Persist per-class next-id counters to disk."""
    try:
        with open(ID_COUNTERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(counters, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ Could not save ID counters: {e}")


def init_next_id_counters():
    """Initialize counters by scanning existing students to avoid reusing IDs.

    The counters are keyed by the numeric class prefix (as string). For each
    class we set next = max(existing) + 1, or 1 if none found.
    """
    counters = load_id_counters()
    students = load_students()

    # Ensure every class in CLASS_ID_MAP has an entry
    for class_name, class_num in CLASS_ID_MAP.items():
        key = str(class_num)
        max_seen = 0
        prefix = f"APY{class_num}"
        for s in students:
            sid = s.get('student_id', '')
            if not sid or not sid.startswith(prefix):
                continue
            # Remove prefix and any desk suffix
            tail = sid[len(prefix):].split('-', 1)[0]
            if tail.isdigit():
                try:
                    val = int(tail)
                    if val > max_seen:
                        max_seen = val
                except Exception:
                    pass

        counters[key] = max_seen + 1 if max_seen >= 0 else 1

    # Also ensure unknown/default class prefix exists
    if '99' not in counters:
        counters['99'] = counters.get('99', 1)

    save_id_counters(counters)

def find_student_by_card_uid(card_uid):
    """Find student by any of their card UIDs"""
    students = load_students()
    for student in students:
        if card_uid in student.get('card_uids', []):
            return student
    return None

def generate_student_uid(class_name, desk_id=None):
    """
    Generate UID in format: APY{Class ID Prefix}{Sequential Number}[-{Desk ID}]
    Uses CLASS_ID_MAP for consistent ID prefixes based on class/level
    For Computer Class: APY5001-A5 (includes desk ID)
    For other classes: APY0001 (no desk ID)
    """
    # Get the ID prefix from the mapping
    class_num = CLASS_ID_MAP.get(class_name, 99)  # Default to 99 if not found
    
    if class_num == 99:
        print(f"⚠️ Warning: Class '{class_name}' not found in CLASS_ID_MAP")
        # Try to extract first digit if it exists
        match = re.search(r'\d+', class_name) if class_name else None
        class_num = match.group() if match else "99"
    
    # Use persistent per-class counters to avoid reusing IDs when students are deleted
    counters = load_id_counters()
    key = str(class_num)

    # If counter missing, compute a safe starting point by scanning existing IDs
    if key not in counters:
        max_seen = 0
        prefix = f"APY{class_num}"
        for s in load_students():
            sid = s.get('student_id', '')
            if not sid or not sid.startswith(prefix):
                continue
            tail = sid[len(prefix):].split('-', 1)[0]
            if tail.isdigit():
                try:
                    val = int(tail)
                    if val > max_seen:
                        max_seen = val
                except Exception:
                    pass
        counters[key] = max_seen + 1

    seq = counters.get(key, 1)

    # Persist increment for next use
    counters[key] = seq + 1
    save_id_counters(counters)

    # Generate UID: APY{Class ID Prefix}{Seq padded to 3 digits}
    uid = f"APY{class_num}{str(seq).zfill(3)}"
    
    # For Computer Class, append desk ID
    if desk_id and schedule_is_computer_class(class_name):
        uid = f"{uid}-{desk_id}"
    
    return uid

def save_student_photo(file, student_id):
    """
    Save uploaded photo for a student
    Returns the photo filename or None if saving failed
    """
    try:
        if not file or file.filename == '':
            return None
        
        # Get file extension
        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        
        if ext not in ALLOWED_PHOTO_EXTENSIONS:
            print(f"⚠️ Invalid photo extension: {ext}")
            return None
        
        if file.content_length and file.content_length > MAX_PHOTO_SIZE:
            print(f"⚠️ Photo too large: {file.content_length} bytes")
            return None
        
        # Create filename: student_id.extension
        photo_filename = f"{student_id}.{ext}"
        photo_path = os.path.join(PHOTOS_DIR_PATH, photo_filename)

        # Save the file as raw bytes to preserve original resolution/quality
        try:
            file.stream.seek(0)
        except Exception:
            pass

        try:
            data = file.read()
            with open(photo_path, 'wb') as out_f:
                out_f.write(data)
            print(f"✅ Photo saved: {photo_path}")
            return photo_filename
        except Exception as e:
            print(f"❌ Error writing photo bytes: {e}")
            return None
    except Exception as e:
        print(f"❌ Error saving photo: {e}")
        return None

def get_default_photo(gender):
    """Return default photo filename based on a variety of gender values."""
    if not gender:
        return DEFAULT_PHOTOS.get('male')
    g = gender.strip().lower()
    if g in ('male', 'boy', 'ប្រុស'):
        return DEFAULT_PHOTOS.get('male')
    if g in ('female', 'girl', 'ស្រី'):
        return DEFAULT_PHOTOS.get('female')
    # Treat common monk/novice keywords (including Khmer) as monk
    if 'monk' in g or g in ('monk', 'monkdefalt', 'monkdefaltprofile', 'បព្វជិត'):
        return DEFAULT_PHOTOS.get('monk')
    return DEFAULT_PHOTOS.get('male')

def find_student_by_name_sex(name, sex):
    """Find student by name and sex (to detect same student registering for multiple classes)"""
    students = load_students()
    for student in students:
        if student.get('name', '').strip() == name.strip() and student.get('sex', '') == sex:
            return student
    return None

# ======================== TIME WINDOW VALIDATION FUNCTIONS ========================

def get_student_schedules(student):
    """Get all schedules for a student (handles both old and new formats)"""
    schedules = student.get('schedules', [])
    if not schedules and 'schedule' in student:
        schedules = [student['schedule']]
    return schedules

def get_primary_schedule(student):
    """Get the primary schedule for a student"""
    schedules = get_student_schedules(student)
    return schedules[0] if schedules else None

def time_to_minutes(time_str):
    """Convert HH:MM time string to minutes since midnight"""
    try:
        hours, minutes = map(int, time_str.split(':'))
        return hours * 60 + minutes
    except:
        return -1

def minutes_to_time(minutes):
    """Convert minutes since midnight to HH:MM string"""
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours:02d}:{mins:02d}"

def get_current_session_for_checkin(student, now=None):
    """
    Determine which session a student should be checking into based on current time.
    Returns: (session_number, schedule_key) or (None, None) if not within any session window
    """
    if now is None:
        now = datetime.now()
    
    current_time_str = now.strftime('%H:%M')
    current_minutes = time_to_minutes(current_time_str)
    
    schedules = get_student_schedules(student)
    
    for schedule in schedules:
        if schedule not in CLASS_CONFIG:
            continue
        
        config = CLASS_CONFIG[schedule]
        
        # Check each session's check-in window
        for session_num in range(1, config.get('sessions', 1) + 1):
            session_key = f'session_{session_num}'
            if session_key not in config.get('schedule_times', {}):
                continue
            
            session_times = config['schedule_times'][session_key]
            checkin_start = time_to_minutes(session_times['checkin_start'])
            checkin_end = time_to_minutes(session_times['checkin_end'])
            
            # Check if current time is within check-in window
            if checkin_start <= current_minutes <= checkin_end:
                return (session_num, schedule)
    
    return (None, None)

def validate_checkin_window(schedule, session_num, now=None):
    """
    Validate if current time is within the allowed check-in window for a session.
    Returns: {'valid': bool, 'status': str, 'time_remaining': int_minutes}
    """
    if now is None:
        now = datetime.now()
    
    if schedule not in CLASS_CONFIG:
        return {'valid': False, 'status': 'INVALID_SCHEDULE', 'time_remaining': 0}
    
    config = CLASS_CONFIG[schedule]
    session_key = f'session_{session_num}'
    
    if session_key not in config.get('schedule_times', {}):
        return {'valid': False, 'status': 'INVALID_SESSION', 'time_remaining': 0}
    
    session_times = config['schedule_times'][session_key]
    current_time_str = now.strftime('%H:%M')
    current_minutes = time_to_minutes(current_time_str)
    
    checkin_start = time_to_minutes(session_times['checkin_start'])
    checkin_end = time_to_minutes(session_times['checkin_end'])
    grace = GRACE_WINDOW_MINUTES

    if current_minutes < checkin_start:
        minutes_early = checkin_start - current_minutes
        return {
            'valid': False,
            'status': 'TOO_EARLY',
            'time_remaining': minutes_early,
            'message': f'Check-in starts at {session_times["checkin_start"]} ({minutes_early}min away)'
        }
    elif current_minutes <= checkin_end:
        # Within scheduled window
        minutes_since_start = current_minutes - checkin_start
        return {
            'valid': True,
            'status': 'ON_TIME',
            'time_remaining': checkin_end - current_minutes,
            'minutes_since_start': minutes_since_start
        }
    elif current_minutes <= checkin_end + grace:
        # Within grace window after scheduled end — allow as LATE
        minutes_late = current_minutes - checkin_end
        grace_until = minutes_to_time(checkin_end + grace)
        return {
            'valid': True,
            'status': 'LATE',
            'time_remaining': 0,
            'minutes_late': minutes_late,
            'message': f'Late — within {grace}min grace (until {grace_until})'
        }
    else:
        # Outside grace window
        return {
            'valid': False,
            'status': 'TOO_LATE',
            'time_remaining': 0,
            'message': f'Check-in ended at {session_times["checkin_end"]}'
        }

def get_attendance_status(schedule, session_num, check_in_time_str, now=None):
    """
    Determine attendance status based on check-in time vs schedule window.
    Returns: {'status': 'EARLY'|'ON_TIME'|'LATE', 'message': str}
    """
    if now is None:
        now = datetime.now()
    
    if schedule not in CLASS_CONFIG:
        return {'status': 'UNKNOWN', 'message': 'Invalid schedule'}
    
    config = CLASS_CONFIG[schedule]
    session_key = f'session_{session_num}'
    
    if session_key not in config.get('schedule_times', {}):
        return {'status': 'UNKNOWN', 'message': 'Invalid session'}
    
    session_times = config['schedule_times'][session_key]
    checkin_start_str = session_times['checkin_start']
    checkin_end_str = session_times['checkin_end']
    checkin_start = time_to_minutes(checkin_start_str)
    checkin_end = time_to_minutes(checkin_end_str)
    checkin_minutes = time_to_minutes(check_in_time_str)
    grace = GRACE_WINDOW_MINUTES

    if checkin_minutes < checkin_start:
        return {
            'status': 'EARLY',
            'message': f'✅ Early check-in (before {checkin_start_str})'
        }
    elif checkin_minutes <= checkin_end:
        return {
            'status': 'ON_TIME',
            'message': f'✅ On-time check-in'
        }
    elif checkin_minutes <= checkin_end + grace:
        minutes_late = checkin_minutes - checkin_end
        return {
            'status': 'LATE',
            'message': f'✅ Late check-in ({minutes_late}min late, within {grace}min grace)'
        }
    else:
        return {
            'status': 'LATE_EXCEEDED',
            'message': f'❌ Check-in outside allowed window (after grace period)'
        }


def infer_weekdays_from_schedule(schedule_name):
    """Heuristic to infer which weekdays a schedule runs on.
    Returns a set of weekday integers where Monday==0 ... Sunday==6.
    Defaults to Monday-Friday when unsure.
    """
    if not schedule_name:
        return set(range(0, 5))
    s = schedule_name.lower()
    # Khmer keywords
    if 'សៅរ៍' in s and 'អាទិត្យ' in s:
        return {5, 6}
    if 'សៅរ៍' in s and 'ចន្ទ' not in s:
        return {5}
    if 'អាទិត្យ' in s and 'សៅរ៍' not in s:
        return {6}
    # Common phrasing
    if 'sat' in s or 'saturday' in s:
        if 'sun' in s or 'sunday' in s:
            return {5, 6}
        return {5}
    if 'sun' in s or 'sunday' in s:
        return {6}
    if 'mon' in s and 'fri' in s:
        return set(range(0, 5))
    if 'weekend' in s:
        return {5, 6}
    # Default to weekdays Monday-Friday
    return set(range(0, 5))


def calculate_absences(student_id, start_date, end_date, attendance_csv=ATTENDANCE_LOG_CSV):
    """Calculate number of absent sessions for a student between two dates (inclusive).

    Rules implemented:
      - Do not count dates before the student's registration (inferred from students.json or first log entry).
      - Only count dates that fall on the schedule's meeting weekdays (heuristic).
      - If the entire section (schedule) has no scans for a date, skip that date (unusual day off).
      - Sessions are counted separately (In1/Out1 and In2/Out2). Missing an In for a session counts as absent for that session.

    Returns: dict { 'student_id': str, 'absent_sessions': int, 'details': {date: [missing_session_nums]} }
    """
    # Normalize dates
    try:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
    except Exception:
        raise ValueError('start_date and end_date must be YYYY-MM-DD')

    students = load_students()
    student = next((s for s in students if s.get('student_id') == student_id), None)
    if not student:
        raise ValueError('student_id not found')

    # Determine registration date: prefer explicit field, else first attendance log for this student
    reg_date = None
    if student.get('registered_at'):
        try:
            reg_date = datetime.strptime(student['registered_at'][:10], '%Y-%m-%d').date()
        except Exception:
            reg_date = None

    # Build student->schedule mapping (primary schedule)
    mapping = {}
    for s in students:
        if 'schedule' in s:
            mapping[s['student_id']] = s['schedule']
        elif 'schedules' in s and s['schedules']:
            mapping[s['student_id']] = s['schedules'][0]

    primary_schedule = mapping.get(student_id)
    if not primary_schedule:
        # no schedule - zero absences
        return {'student_id': student_id, 'absent_sessions': 0, 'details': {}}

    # Parse attendance log to find earliest date for this student and to index by date
    logs_by_date = {}
    section_has_scan = {}  # date -> bool: whether any student in this schedule scanned
    if os.path.exists(attendance_csv):
        with open(attendance_csv, 'r', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                d = row.get('Date', '')
                if not d:
                    continue
                try:
                    date_obj = datetime.strptime(d, '%Y-%m-%d').date()
                except Exception:
                    continue

                # Map rows by date and student
                if date_obj not in logs_by_date:
                    logs_by_date[date_obj] = {}

                sid = row.get('StudentID')
                logs_by_date[date_obj][sid] = row

                # Mark section scan if this row belongs to this schedule and has any In/Out
                sid_schedule = mapping.get(sid)
                if sid_schedule == primary_schedule:
                    has_scan = bool(row.get('In1') or row.get('In2') or row.get('Out1') or row.get('Out2'))
                    if date_obj not in section_has_scan:
                        section_has_scan[date_obj] = has_scan
                    else:
                        section_has_scan[date_obj] = section_has_scan[date_obj] or has_scan

                # If reg_date not set and this row is for the student, use it as earliest
                if sid == student_id and reg_date is None:
                    try:
                        reg_date = date_obj if reg_date is None else min(reg_date, date_obj)
                    except Exception:
                        reg_date = date_obj

    # If still no reg_date, assume start_date
    if reg_date is None:
        reg_date = start_dt

    # Effective start is max(start_dt, reg_date)
    effective_start = max(start_dt, reg_date)

    # Determine which weekdays the schedule runs on
    weekdays = infer_weekdays_from_schedule(primary_schedule)

    absent_sessions = 0
    details = {}

    cur = effective_start
    while cur <= end_dt:
        # Skip dates before registration
        if cur < reg_date:
            cur = cur + timedelta(days=1)
            continue

        # Weekday filter
        if cur.weekday() not in weekdays:
            cur = cur + timedelta(days=1)
            continue

        # Section-wide off-day: if no scans from anyone in this schedule, skip
        if cur in section_has_scan and not section_has_scan[cur]:
            cur = cur + timedelta(days=1)
            continue

        # Determine sessions count for this schedule
        cfg = CLASS_CONFIG.get(primary_schedule, {})
        sessions = cfg.get('sessions', 1)

        # Lookup student's row for this date
        row = logs_by_date.get(cur, {}).get(student_id)

        missing = []
        for sn in range(1, sessions + 1):
            in_field = f'In{sn}'
            # If no row or empty In field -> absent for that session
            if not row or not (row.get(in_field) and row.get(in_field).strip()):
                missing.append(sn)

        if missing:
            absent_sessions += len(missing)
            details[cur.strftime('%Y-%m-%d')] = missing

        cur = cur + timedelta(days=1)

    return {'student_id': student_id, 'absent_sessions': absent_sessions, 'details': details}


def is_browser_request():
    """Return True if the incoming request looks like it's from a browser (HTML Accept or common UA)."""
    try:
        accept = (request.headers.get('Accept') or '').lower()
        ua = (request.headers.get('User-Agent') or '').lower()
        if 'text/html' in accept:
            return True
        if ua.startswith('mozilla') or 'chrome' in ua or 'safari' in ua or 'edge' in ua:
            return True
    except Exception:
        pass
    return False

def validate_checkout_window(schedule, session_num, now=None):
    """
    Validate if current time is within the allowed check-out window for a session.
    Returns: {'valid': bool, 'status': str, 'message': str}
    """
    if now is None:
        now = datetime.now()
    
    if schedule not in CLASS_CONFIG:
        return {'valid': False, 'status': 'INVALID_SCHEDULE', 'message': 'Invalid schedule'}
    
    config = CLASS_CONFIG[schedule]
    session_key = f'session_{session_num}'
    
    if session_key not in config.get('schedule_times', {}):
        return {'valid': False, 'status': 'INVALID_SESSION', 'message': 'Invalid session'}
    
    session_times = config['schedule_times'][session_key]
    current_time_str = now.strftime('%H:%M')
    current_minutes = time_to_minutes(current_time_str)
    
    checkout_start = time_to_minutes(session_times['checkout_start'])
    checkout_end = time_to_minutes(session_times['checkout_end'])
    
    if current_minutes < checkout_start:
        minutes_early = checkout_start - current_minutes
        return {
            'valid': False,
            'status': 'TOO_EARLY_CHECKOUT',
            'message': f'⏱️ Checkout opens at {session_times["checkout_start"]} ({minutes_early}min away)'
        }
    elif current_minutes > checkout_end:
        return {
            'valid': True,
            'status': 'LATE_CHECKOUT',
            'message': f'⏱️ Late checkout (after {session_times["checkout_end"]})'
        }
    else:
        # Within window
        return {
            'valid': True,
            'status': 'ON_TIME_CHECKOUT',
            'message': f'✅ On-time checkout'
        }

init_student_data()
# Initialize persistent next-id counters (scan existing students)
try:
    init_next_id_counters()
except Exception as e:
    print(f"⚠️ Could not initialize ID counters: {e}")

# ======================== CORE ROUTES ========================

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')


@app.route('/')
def home():
    """Render the new homepage."""
    # Collect slide images from resource/Slide if available
    slides_dir = os.path.join(PROJECT_ROOT, 'resource', 'Slide')
    slides = []
    try:
        if os.path.exists(slides_dir) and os.path.isdir(slides_dir):
            for fname in sorted(os.listdir(slides_dir)):
                if fname.lower().endswith(('.jpg', '.jpeg', '.png', '.webp', '.gif')):
                    slides.append(f"/resource/Slide/{fname}")
    except Exception as e:
        print(f"⚠️ Error listing slides: {e}")

    # Logo path (spaces encoded)
    logo_path = '/resource/icon&symbol/School%20Logo.jpg'

    return render_template('home.html', slides=slides, logo_url=logo_path), "Hello from hte Flask Mobile Server!"

def start_mdns():
    port = 8080

    # Get your hotspot IP address dynamically
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        ip_address = s.getsockname()[0]
    except Exception:
        ip_address = "10.147.174.100"  # Your hotspot fallback IP
    finally:
        s.close()

    # The clean setup that satisfies python-zeroconf validation rules:
    info = ServiceInfo(
        type_="_http._tcp.local.",
        name="nfc._http._tcp.local.",  # Combine your name + service type
        addresses=[socket.inet_aton(ip_address)],
        port=port,
        properties={},
        server="nfc.local.",  # The host target name
    )

    zeroconf = Zeroconf()
    zeroconf.register_service(info)
    print(f"📡 mDNS active! Try connecting to: http://nfc.local:{port}")
	
@app.route('/checkin_page')
def checkin_page():
    # Attempt to embed the waiting GIF as a data URI to avoid any serving issues
    try:
        gif_path = os.path.join(PROJECT_ROOT, 'resource', 'scanning1.gif')
        if os.path.exists(gif_path) and os.path.getsize(gif_path) < (2 * 1024 * 1024):
            import base64
            with open(gif_path, 'rb') as f:
                data = f.read()
            waiting_gif_data = 'data:image/gif;base64,' + base64.b64encode(data).decode('ascii')
            return render_template('checkin.html', waiting_gif_data=waiting_gif_data)
    except Exception as e:
        print(f"⚠️ Could not embed waiting GIF: {e}")

    return render_template('checkin.html', waiting_gif_data=None)


@app.route('/checkin?uid=<card_uid>')
@app.route('/checkin/<card_uid>')
def checkin_redirect(card_uid):
        """Handle legacy kiosk-style checkin URLs like /checkin=10323D55
        and redirect the browser to the student's report after a short delay.
        """
        try:
                card_uid = (card_uid or '').strip()
                student = find_student_by_card_uid(card_uid)
                student_id = student.get('student_id') if student else request.args.get('student_id') or 'APY0001'
                redirect_url = url_for('student_report', student_id=student_id)

                # Simple HTML page with meta-refresh and JS fallback after 3 seconds
                html = f'''
                <!doctype html>
                <html lang="en">
                    <head>
                        <meta charset="utf-8" />
                        <meta http-equiv="refresh" content="3;url={redirect_url}" />
                        <title>Redirecting...</title>
                        <style>body{{font-family:system-ui,Segoe UI,Roboto,Arial;margin:40px;color:#333}}</style>
                    </head>
                    <body>
                        <h3>Redirecting to student report: {student_id}</h3>
                        <p>If your browser does not redirect automatically, <a href="{redirect_url}">click here</a>.</p>
                        <script>setTimeout(function(){{window.location.href = "{redirect_url}";}}, 5000);</script>
                    </body>
                </html>
                '''
                return html
        except Exception as e:
                print(f"⚠️ checkin_redirect error: {e}")
                return redirect(url_for('checkin_page'))

@app.route('/audio_test')
def audio_test():
    """Audio system diagnostics page - for testing if sounds are working"""
    return render_template('audio_test.html')

@app.route('/simulator')
def simulator():
    return render_template('simulator.html')

@app.route('/card_management')
def card_management():
    return render_template('card_management.html')

@app.route('/attendance_report')
def attendance_report():
    return render_template('attendance_report.html')

@app.route('/api/last_scan')
def get_last_scan():
    global DUPLICATE_ERROR_TIMESTAMP
    
    # Make a copy to return
    scan_data = dict(LAST_SCAN_DATA)
    
    # If there's a duplicate error, only return it if it's recent (within 15 seconds)
    # This prevents stale duplicate errors from showing repeatedly
    if scan_data.get('status') == 'DUPLICATE_CARD' and DUPLICATE_ERROR_TIMESTAMP:
        time_diff = (datetime.now() - DUPLICATE_ERROR_TIMESTAMP).total_seconds()
        if time_diff > 15:  # Clear error if older than 15 seconds
            scan_data['status'] = None
            scan_data['message'] = None
    
    return jsonify(scan_data)

@app.route('/api/clear_duplicate_error', methods=['POST'])
def clear_duplicate_error():
    """Clear the duplicate error state to prevent it from showing repeatedly"""
    global DUPLICATE_ERROR_TIMESTAMP, LAST_SCAN_DATA
    
    DUPLICATE_ERROR_TIMESTAMP = None
    # Only clear if it was a duplicate card error
    if LAST_SCAN_DATA.get('status') == 'DUPLICATE_CARD':
        LAST_SCAN_DATA = {"uid": None, "name": None, "status": None, "timestamp": None, "message": None}
    
    return jsonify({"status": "cleared"}), 200

@app.route('/api/checkin_stream')
def checkin_stream():
    """Server-Sent Events endpoint for real-time check-in updates"""
    def generate():
        last_sent = None
        stale_count = 0
        # Snapshot the current LAST_SCAN_DATA at connect time so we don't replay
        # the most recent check-in to newly-connecting clients. Only send when
        # LAST_SCAN_DATA changes after the connection was established.
        initial_snapshot = dict(LAST_SCAN_DATA)
        
        while True:
            # Send heartbeat if no new data or if the last-known data is the same
            # as the snapshot at connect time (prevents replaying the last event).
            if LAST_SCAN_DATA == last_sent or (last_sent is None and LAST_SCAN_DATA == initial_snapshot):
                stale_count += 1
                yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
                
                # Disconnect if client inactive for 30 seconds
                if stale_count > 30:
                    break
            else:
                stale_count = 0
                last_sent = dict(LAST_SCAN_DATA)
                event_data = {
                    'type': 'checkin',
                    'uid': LAST_SCAN_DATA.get('uid'),
                    'name': LAST_SCAN_DATA.get('name'),
                    'status': LAST_SCAN_DATA.get('status'),
                    'message': LAST_SCAN_DATA.get('message'),
                    'timestamp': LAST_SCAN_DATA.get('timestamp'),
                    'slot': LAST_SCAN_DATA.get('slot'),
                    'sound': LAST_SCAN_DATA.get('sound')
                }

                # Debug logging: if a sound is attached, print before and after sending
                try:
                    if event_data.get('sound'):
                        print(f"[SSE DEBUG] About to send checkin event with sound='{event_data.get('sound')}' uid={event_data.get('uid')} name={event_data.get('name')}")
                except Exception as _e:
                    print(f"[SSE DEBUG] Pre-send logging failed: {_e}")

                payload = f"data: {json.dumps(event_data)}\n\n"
                yield payload

                try:
                    if event_data.get('sound'):
                        print(f"[SSE DEBUG] Sent checkin event with sound='{event_data.get('sound')}' uid={event_data.get('uid')} name={event_data.get('name')}")
                except Exception as _e:
                    print(f"[SSE DEBUG] Post-send logging failed: {_e}")

                # Also add to recent checkins
                if LAST_SCAN_DATA.get('uid'):
                    RECENT_CHECKINS.append(event_data)
            
            time.sleep(0.5)  # Check every 500ms for new data
    
    return Response(generate(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no',
        'Connection': 'keep-alive'
    })

@app.route('/api/recent_checkins')
def get_recent_checkins():
    """Get recent check-ins for initial page load"""
    return jsonify(list(RECENT_CHECKINS))

@app.route('/api/registration_stream')
def registration_stream():
    """Server-Sent Events endpoint for real-time registration updates"""
    def generate():
        last_sent = None
        stale_count = 0
        
        while True:
            # Send heartbeat if no new data
            if LAST_REGISTRATION == last_sent:
                stale_count += 1
                yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
                
                # Disconnect if client inactive for 30 seconds
                if stale_count > 30:
                    break
            else:
                stale_count = 0
                last_sent = dict(LAST_REGISTRATION) if LAST_REGISTRATION else None
                if LAST_REGISTRATION:
                    event_data = {
                        'type': 'registration',
                        'uid': LAST_REGISTRATION.get('uid'),
                        'name': LAST_REGISTRATION.get('name'),
                        'sex': LAST_REGISTRATION.get('sex'),
                        'schedule': LAST_REGISTRATION.get('schedule'),
                        'message': LAST_REGISTRATION.get('message'),
                        'timestamp': LAST_REGISTRATION.get('timestamp')
                    }
                    yield f"data: {json.dumps(event_data)}\n\n"
            
            time.sleep(0.5)  # Check every 500ms for new data
    
    return Response(generate(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no',
        'Connection': 'keep-alive'
    })

@app.route('/api/student_status/<uid>')
def get_student_status(uid):
    """Fetch current student status from attendance log (for kiosk to verify completion)"""
    uid = uid.upper()
    today = datetime.now().strftime('%Y-%m-%d')
    
    try:
        with open(ATTENDANCE_LOG_CSV, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row['Date'] == today and row['UID'] == uid:
                    # Return the current attendance record
                    is_present = (row['In1'] and row['Out1'] and 
                                (not CLASS_CONFIG.get(f"Class {uid[3]}", {}).get("sessions", 2) == 2 or 
                                 (row['In2'] and row['Out2'])))
                    
                    return jsonify({
                        "uid": uid,
                        "name": row['Name'],
                        "in1": row['In1'],
                        "out1": row['Out1'],
                        "in2": row['In2'],
                        "out2": row['Out2'],
                        "status": row['Status'],
                        "is_complete": row['Status'] == 'PRESENT'
                    }), 200
        
        # Not found for today
        return jsonify({"error": "No record for today", "uid": uid}), 404
    except Exception as e:
        print(f"❌ Student Status Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/class_config')
def get_class_config():
    """Get class schedule configuration (sessions per class)"""
    return jsonify(CLASS_CONFIG)

@app.route('/api/students')
def get_students():
    try:
        students_data = load_students()
        students = []
        for student in students_data:
            # Handle both old (schedule) and new (schedules) formats
            schedules = student.get('schedules', [])
            if not schedules and 'schedule' in student:
                schedules = [student['schedule']]
            
            students.append({
                'ID': student.get('student_id', 'Unknown'),
                'Name': student.get('name', 'Unknown'),
                'sex': student.get('sex', 'Unknown'),
                'schedule': schedules[0] if schedules else 'Unknown',  # Primary schedule
                'schedules': schedules,  # All schedules
                'card_uids': student.get('card_uids', []),
                'desk_ids': student.get('desk_ids', {}),
                'photo': student.get('photo') if student.get('photo') else get_default_photo(student.get('sex'))
            })
        return jsonify(students)
    except Exception as e:
        print(f"❌ Students API Error: {e}")
        return jsonify([]), 500

@app.route('/api/student/<student_id>')
def get_student(student_id):
    """Get student data by student_id, including photo information"""
    try:
        students = load_students()
        for student in students:
            if student.get('student_id') == student_id:
                # Determine photo filename (use default if missing)
                photo_filename = student.get('photo') if student.get('photo') else get_default_photo(student.get('sex'))

                return jsonify({
                    'student_id': student.get('student_id'),
                    'name': student.get('name'),
                    'sex': student.get('sex'),
                    'photo': photo_filename,
                    'schedules': student.get('schedules', []),
                    'card_uids': student.get('card_uids', [])
                }), 200
        
        # Student not found
        return jsonify({"error": "Student not found"}), 404
    except Exception as e:
        print(f"❌ Student API Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/student_update/<student_id>', methods=['POST'])
def update_student_api(student_id):
    """Update student's profile fields (name, dob) and optional photo upload.
    Accepts multipart/form-data with optional 'photo' file and form fields 'name' and 'dob'.
    """
    try:
        students = load_students()
        student = next((s for s in students if s.get('student_id') == student_id), None)
        if not student:
            return jsonify({'error': 'Student not found'}), 404

        changed = {}

        # Accept JSON body as well
        data = {}
        if request.is_json:
            data = request.get_json(silent=True) or {}
        else:
            data = request.form or {}

        name = data.get('name')
        dob = data.get('dob')

        if name is not None:
            old = student.get('name', '')
            student['name'] = name
            changed['name'] = {'old': old, 'new': name}

        if dob is not None:
            old = student.get('dob', '')
            student['dob'] = dob
            changed['dob'] = {'old': old, 'new': dob}

        # Handle photo upload
        if 'photo' in request.files:
            f = request.files['photo']
            if f and f.filename:
                filename = secure_filename(f.filename)
                ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
                if ext not in ALLOWED_PHOTO_EXTENSIONS:
                    return jsonify({'error': 'Invalid photo extension'}), 400
                # read bytes to enforce size limit
                data_bytes = f.read()
                if len(data_bytes) > MAX_PHOTO_SIZE:
                    return jsonify({'error': 'Photo too large'}), 413
                # generate stable filename per student to avoid collisions
                out_name = f"{student_id}.{ext}"
                out_path = os.path.join(PHOTOS_DIR_PATH, out_name)
                try:
                    with open(out_path, 'wb') as of:
                        of.write(data_bytes)
                    student['photo'] = out_name
                    changed['photo'] = out_name
                except Exception as e:
                    return jsonify({'error': f'Could not save photo: {e}'}), 500

        # Persist changes if any
        if changed:
            save_students(students)
            # Simple audit entry (no admin tracking here)
            try:
                append_audit_entry('web', student_id, datetime.now().strftime('%Y-%m-%d'), 'profile_update', json.dumps(changed), '')
            except Exception:
                pass

        return jsonify({'ok': True, 'changed': changed, 'student': {'student_id': student_id, 'name': student.get('name'), 'dob': student.get('dob'), 'photo': student.get('photo')}}), 200
    except Exception as e:
        print(f"❌ Student Update Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/student_report/<student_id>')
def student_report(student_id):
    """Render a simple report page for an individual student."""
    try:
        students = load_students()
        student = next((s for s in students if s.get('student_id') == student_id), None)
    except Exception as e:
        print(f"❌ Student Report Load Error: {e}")
        student = None

    if not student:
        return render_template('student_report.html', student=None, error='Student not found'), 404

    # Determine photo filename (use default if missing)
    photo_filename = student.get('photo') if student.get('photo') else get_default_photo(student.get('sex'))

    # Load attendance records for this student
    records = []
    try:
        with open(ATTENDANCE_LOG_CSV, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('StudentID') == student_id:
                    records.append(row)
    except Exception as e:
        print(f"⚠️ Could not load attendance for {student_id}: {e}")

    # Build calendar view for a given month/year (query params optional)
    try:
        now = datetime.now()
        year = int(request.args.get('year', now.year))
        month = int(request.args.get('month', now.month))
    except Exception:
        year = datetime.now().year
        month = datetime.now().month

    # Map day -> status summary for that day
    day_status = {}  # e.g., {1: 'present', 2: 'late', 3: 'absent', 4: 'permission'}
    counts = {'present': 0, 'late': 0, 'absent': 0, 'permission': 0, 'leave': 0}

    # Prepare fast lookup of records by date
    rec_by_date = {}
    for r in records:
        d = r.get('Date', '')
        try:
            if d.startswith(f"{year}-{str(month).zfill(2)}-"):
                day = int(d.split('-')[-1])
                rec_by_date[day] = r
        except Exception:
            continue

    # Determine registration date for this student to avoid counting absences before enrollment
    reg_date = None
    if student.get('registered_at'):
        try:
            reg_date = datetime.strptime(student['registered_at'], '%Y-%m-%d').date()
        except Exception:
            reg_date = None

    # If no explicit registration date, infer from earliest attendance record for the student
    if not reg_date and records:
        try:
            student_dates = [datetime.strptime(r.get('Date'), '%Y-%m-%d').date() for r in records if r.get('Date')]
            if student_dates:
                reg_date = min(student_dates)
        except Exception:
            reg_date = None

    # Default to today's date if still unknown (so we don't count past days as absent)
    if not reg_date:
        reg_date = datetime.now().date()

    # Determine student schedule for attendance status calculations
    student_schedule = get_primary_schedule(student) or (student.get('schedule') if student.get('schedule') else None)

    month_range = calendar.monthrange(year, month)[1]
    for day in range(1, month_range + 1):
        try:
            this_date = datetime(year, month, day).date()
        except Exception:
            day_status[day] = 'absent'
            counts['absent'] += 1
            continue

        # Skip days before the student registered
        if this_date < reg_date:
            day_status[day] = 'not_enrolled'
            continue

        # Skip future days (incoming days) as neutral so they are not counted as absences
        if this_date > datetime.now().date():
            day_status[day] = 'future'
            continue

        r = rec_by_date.get(day)
        if not r:
            day_status[day] = 'absent'
            counts['absent'] += 1
            continue

        status_field = (r.get('Status') or '').upper()
        # Permission/excused markers
        if status_field in ('EXCUSED', 'PERMISSION', 'PERM'):
            day_status[day] = 'permission'
            counts['permission'] += 1
            continue

        # If there's a recorded In1 or In2, evaluate presence/late/leave
        checkin_time = r.get('In1') or r.get('In2') or ''
        checkout_exists = bool(r.get('Out1') or r.get('Out2'))
        if checkin_time:
            # Determine which session this check-in best belongs to
            session_num = 1
            if student_schedule:
                cfg = CLASS_CONFIG.get(student_schedule, {})
                for sn in range(1, cfg.get('sessions', 1) + 1):
                    sk = f'session_{sn}'
                    stimes = cfg.get('schedule_times', {}).get(sk)
                    if not stimes:
                        continue
                    st_end = time_to_minutes(stimes.get('checkin_end'))
                    ct = time_to_minutes(checkin_time)
                    if ct <= st_end + GRACE_WINDOW_MINUTES:
                        session_num = sn
                        break

            # Compute minutes late relative to session start
            late_threshold_minutes = 20
            if student_schedule:
                cfg = CLASS_CONFIG.get(student_schedule, {})
                sk = f'session_{session_num}'
                session_times = cfg.get('schedule_times', {}).get(sk, {})
                session_start = time_to_minutes(session_times.get('checkin_start', '00:00'))
            else:
                session_start = 0

            ct_minutes = time_to_minutes(checkin_time)
            minutes_after_start = ct_minutes - session_start

            # If check-in is more than late_threshold_minutes after session start, count as late
            if minutes_after_start > late_threshold_minutes:
                day_status[day] = 'late'
                counts['late'] += 1
            else:
                # If there's a checkout recorded, consider fully present; else mark as leave
                if checkout_exists:
                    day_status[day] = 'present'
                    counts['present'] += 1
                else:
                    day_status[day] = 'leave'
                    counts.setdefault('leave', 0)
                    counts['leave'] += 1
        else:
            # No checkin times and no explicit In recorded
            if checkout_exists:
                # Checkout without checkin — treat as present
                day_status[day] = 'present'
                counts['present'] += 1
            else:
                day_status[day] = 'absent'
                counts['absent'] += 1

    # Build month calendar weeks for template
    month_weeks = calendar.monthcalendar(year, month)

    return render_template('student_report.html', student=student, photo=photo_filename, records=records, class_config=CLASS_CONFIG, month=month, year=year, weeks=month_weeks, day_status=day_status, counts=counts)


@app.route('/api/student_note/<student_id>', methods=['GET', 'POST'])
def student_note_api(student_id):
    """
    GET: return note content for given student and date (query param `date=YYYY-MM-DD`).
    POST: save note content (JSON {date, note}). Only allows saving for today's date.
    Notes are stored in `notes/<student_id>_<YYYY-MM-DD>.txt`.
    """
    notes_dir = os.path.join(PROJECT_ROOT, 'notes')
    os.makedirs(notes_dir, exist_ok=True)

    notes_file = os.path.join(notes_dir, secure_filename(f"{student_id}.json"))

    def load_notes_map():
        """Return a dict mapping date->note for this student. Performs migration from legacy per-date txt files if needed."""
        # If consolidated JSON exists, use it
        if os.path.exists(notes_file):
            try:
                with open(notes_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data if isinstance(data, dict) else {}
            except Exception:
                return {}

        # Else, look for legacy per-date txt files like {student_id}_YYYY-MM-DD.txt
        notes_map = {}
        pattern = os.path.join(notes_dir, f"{student_id}_*.txt")
        for p in glob.glob(pattern):
            basename = os.path.basename(p)
            try:
                parts = basename.rsplit('.', 1)[0].split('_', 1)
                if len(parts) == 2:
                    date_part = parts[1]
                    with open(p, 'r', encoding='utf-8') as f:
                        notes_map[date_part] = f.read()
            except Exception:
                continue

        # If we found any legacy files, persist consolidated JSON for future use
        if notes_map:
            try:
                with open(notes_file, 'w', encoding='utf-8') as f:
                    json.dump(notes_map, f, ensure_ascii=False, indent=2)
            except Exception:
                pass

        return notes_map

    if request.method == 'GET':
        date = request.args.get('date')
        if not date:
            return jsonify({'error': 'date query parameter required'}), 400

        notes_map = load_notes_map()
        note_text = notes_map.get(date, '')
        # readonly if date is before today
        try:
            d = datetime.strptime(date, '%Y-%m-%d').date()
            readonly = d < datetime.now().date()
        except Exception:
            readonly = True

        return jsonify({'date': date, 'note': note_text, 'readonly': readonly}), 200

    # POST: save note (only for today)
    try:
        data = request.get_json(force=True)
        date = data.get('date')
        note = data.get('note', '')
    except Exception:
        return jsonify({'error': 'Invalid JSON'}), 400

    if not date:
        return jsonify({'error': 'date is required'}), 400

    # Only allow saving for today's date
    try:
        d = datetime.strptime(date, '%Y-%m-%d').date()
        if d != datetime.now().date():
            return jsonify({'error': 'Can only save notes for today'}), 403
    except Exception:
        return jsonify({'error': 'invalid date format, expected YYYY-MM-DD'}), 400

    notes_map = load_notes_map()
    notes_map[date] = note or ''
    try:
        with open(notes_file, 'w', encoding='utf-8') as f:
            json.dump(notes_map, f, ensure_ascii=False, indent=2)
        return jsonify({'ok': True, 'date': date}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/available_desks/<level>')
def get_available_desks(level):
    """
    Get list of available desks for a Computer Class level.
    Returns only desks that are not yet assigned to any student.
    """
    try:
        # All possible desk positions
        desks = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7',
                 'B1', 'B2', 'B3', 'B4', 'B5', 'B6', 'B7',
                 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'T']
        
        # Get all students and find taken desks
        students = load_students()
        taken_desks = set()
        
        for student in students:
            # Handle both old (schedule) and new (schedules) formats
            schedules = student.get('schedules', [])
            if not schedules and 'schedule' in student:
                schedules = [student['schedule']]
            
            student_id = student.get('student_id', '')
            
            # Check if this student is in the requested Computer Class level
            for schedule in schedules:
                if f"Computer Class - {level}" in schedule:
                    # Extract desk ID from student ID (format: APY5001-A5)
                    if '-' in student_id:
                        desk_id = student_id.split('-')[-1]
                        taken_desks.add(desk_id)
                    # Also check desk_ids object for this specific schedule
                    if 'desk_ids' in student and schedule in student['desk_ids']:
                        desk_id = student['desk_ids'][schedule]
                        taken_desks.add(desk_id)
        
        # Return available desks (those not in taken_desks)
        available_desks = [d for d in desks if d not in taken_desks]
        
        return jsonify({
            "level": level,
            "available_desks": available_desks,
            "total_desks": len(desks),
            "taken_desks": list(taken_desks)
        }), 200
    except Exception as e:
        print(f"❌ Available Desks API Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/validate_duplicates', methods=['GET'])
def validate_duplicates():
    """
    Scan entire students.json for duplicate card UIDs across students.
    Useful for admin dashboard to detect data integrity issues.
    
    Returns:
        - status: 'clean' if no duplicates, 'issues_found' if duplicates exist
        - duplicates: list of duplicate cards and which students have them
        - count: number of duplicate issues found
    """
    try:
        students = load_students()
        card_to_students = {}  # Map card UID to list of students who have it
        
        # Build mapping
        for student in students:
            for card_uid in student.get('card_uids', []):
                if card_uid not in card_to_students:
                    card_to_students[card_uid] = []
                card_to_students[card_uid].append({
                    'student_id': student['student_id'],
                    'name': student['name'],
                    'schedule': student['schedule']
                })
        
        # Find duplicates
        duplicates = []
        for card_uid, student_list in card_to_students.items():
            if len(student_list) > 1:
                duplicates.append({
                    'card_uid': card_uid,
                    'students': student_list,
                    'count': len(student_list)
                })
        
        if duplicates:
            print(f"⚠️ DUPLICATE CARDS FOUND: {len(duplicates)} issues")
            for dup in duplicates:
                print(f"   Card {dup['card_uid']} assigned to: {[s['student_id'] for s in dup['students']]}")
        
        return jsonify({
            'status': 'clean' if not duplicates else 'issues_found',
            'duplicates': duplicates,
            'count': len(duplicates),
            'total_students': len(students),
            'total_cards': len(card_to_students)
        }), 200
    except Exception as e:
        print(f"❌ Validation Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/attendance')
def get_attendance():
    try:
        attendance = []
        with open(ATTENDANCE_LOG_CSV, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Convert to dashboard format
                attendance.append({
                    'Timestamp': f"{row['Date']} {row.get('In1', row.get('Out1', '00:00'))}",
                    'UID': row['UID'],
                    'Name': row['Name'],
                    'Status': row['Status'],
                    'CheckinTime': row.get('In1', 'N/A')
                })
        return jsonify(attendance)
    except Exception as e:
        print(f"❌ Attendance API Error: {e}")
        return jsonify([]), 500

@app.route('/api/live_status')
def get_live_status():
    # Check if PN532 is still connected (heartbeat within last 30 seconds)
    now = datetime.now()
    if PN532_STATUS["last_heartbeat"]:
        time_diff = (now - PN532_STATUS["last_heartbeat"]).total_seconds()
        if time_diff > 30:
            PN532_STATUS["status"] = "disconnected"
        elif time_diff > 10:
            PN532_STATUS["status"] = "reconnecting"
    
    # Placeholder for live status - returns basic system info
    return jsonify({
        'recent_scans': [],
        'pending_write': None,
        'pn532_status': PN532_STATUS["status"],
        'esp32_ip': PN532_STATUS["esp32_ip"],
        'last_heartbeat': PN532_STATUS["last_heartbeat"].isoformat() if PN532_STATUS["last_heartbeat"] else None
    })

@app.route('/api/esp32_status', methods=['GET', 'POST'])
def update_esp32_status():
    """ESP32 heartbeat endpoint to report PN532 status"""
    global PN532_STATUS

    if request.method == 'GET':
        # Return the current PN532/ESP32 status for the dashboard/kiosk
        now = datetime.now()
        if PN532_STATUS["last_heartbeat"]:
            time_diff = (now - PN532_STATUS["last_heartbeat"]).total_seconds()
            if time_diff > 30:
                PN532_STATUS["status"] = "disconnected"
            elif time_diff > 10:
                PN532_STATUS["status"] = "reconnecting"

        return jsonify({
            "pn532_connected": PN532_STATUS["status"] == "connected",
            "pn532_status": PN532_STATUS["status"],
            "esp32_ip": PN532_STATUS["esp32_ip"],
            "last_heartbeat": PN532_STATUS["last_heartbeat"].isoformat() if PN532_STATUS["last_heartbeat"] else None
        }), 200

    # POST updates from ESP32
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    pn532_connected = data.get('pn532_connected', False)
    esp32_ip = request.remote_addr

    PN532_STATUS["esp32_ip"] = esp32_ip
    PN532_STATUS["last_heartbeat"] = datetime.now()

    if pn532_connected:
        if PN532_STATUS["status"] != "connected":
            PN532_STATUS["status"] = "connected"
            print(f"✅ PN532 Connected - ESP32 IP: {esp32_ip}")
    else:
        PN532_STATUS["status"] = "disconnected"
        print(f"❌ PN532 Disconnected - ESP32 IP: {esp32_ip}")

    return jsonify({"status": "updated"}), 200


@app.route('/resource/<path:filename>')
def serve_resource_file(filename):
    """Serve files from the local `resource/` folder (for gifs, assets, etc.)."""
    try:
        resource_dir = os.path.join(PROJECT_ROOT, 'resource')
        full_path = os.path.join(resource_dir, filename)
        # Debugging info for resource serving
        try:
            print(f"DEBUG resource_dir={resource_dir}")
            print(f"DEBUG full_path={full_path}")
            print(f"DEBUG abspaths: rd={os.path.abspath(resource_dir)}, fp={os.path.abspath(full_path)}")
            print(f"DEBUG common1={os.path.commonpath([os.path.abspath(resource_dir)])}")
            print(f"DEBUG common2={os.path.commonpath([os.path.abspath(resource_dir), os.path.abspath(full_path)])}")
        except Exception as _e:
            print(f"DEBUG resource debug failed: {_e}")

        if not os.path.commonpath([os.path.abspath(resource_dir)]) == os.path.commonpath([os.path.abspath(resource_dir), os.path.abspath(full_path)]):
            print(f"❌ Resource path traversal attempt: {filename}")
            return ("", 400)

        if os.path.exists(full_path) and os.path.isfile(full_path):
            # Use send_from_directory to ensure correct mimetype and headers
            try:
                return send_from_directory(resource_dir, filename)
            except Exception as e:
                print(f"⚠️ send_from_directory failed for {filename}: {e}, falling back to send_file")
                return send_file(full_path)

        return ("", 404)
    except Exception as e:
        print(f"❌ Resource serve error: {e}")
        return ("", 500)

# ======================== 7-CLASS EXCEL REPORT ========================

@app.route('/register', methods=['GET', 'POST'])
def register():
    global PENDING_REG_DATA, LAST_REGISTRATION, CURRENT_REGISTRATION_SESSION_ID, DUPLICATE_ERROR_TIMESTAMP, PENDING_CONFIRMATIONS
    if request.method == 'GET':
        return render_template('register.html')
    
    # POST: Save form data and generate UID for new student
    class_name = request.form.get('schedule')
    desk_id = request.form.get('desk_id', '')  # For Computer Class only
    full_name = request.form.get('full_name')
    gender = request.form.get('gender')
    # Date of birth parts (day, month, year) - sent as numeric values from the form
    dob_day = request.form.get('dob_day')
    dob_month = request.form.get('dob_month')
    dob_year = request.form.get('dob_year')
    
    print(f"📝 Registration Form Submitted")
    print(f"   Name: {full_name}, Gender: {gender}")
    print(f"   Schedule: {class_name}, Desk: {desk_id}")
    
    generated_uid = generate_student_uid(class_name, desk_id)
    
    # Create new registration session ID to track this registration attempt
    CURRENT_REGISTRATION_SESSION_ID = str(uuid.uuid4())
    # Clear any old duplicate error when starting a new registration
    DUPLICATE_ERROR_TIMESTAMP = None
    
    # Get photo if uploaded and save it immediately (so it survives until card tap)
    photo_file = request.files.get('photo')
    photo_filename = None
    if photo_file and photo_file.filename:
        photo_filename = save_student_photo(photo_file, generated_uid)
        if photo_filename:
            print(f"   Photo saved during registration: {photo_filename}")

    PENDING_REG_DATA = {
        "uid": generated_uid,
        "name": full_name,
        "sex": gender,
        "schedule": class_name,
        "desk_id": desk_id,
        "session_id": CURRENT_REGISTRATION_SESSION_ID,
        "photo": photo_filename,  # store filename (or None)
        "dob_day": dob_day,
        "dob_month": dob_month,
        "dob_year": dob_year
    }

    # Mark this session as pending confirmation (not confirmed yet)
    PENDING_CONFIRMATIONS[CURRENT_REGISTRATION_SESSION_ID] = {
        "card_uid": None,  # Will be filled when card is tapped
        "student_id": generated_uid,
        "name": full_name,
        "sex": gender,
        "schedule": class_name,
        "desk_id": desk_id,
        "photo": photo_filename,  # stored filename (or None)
        "dob_day": dob_day,
        "dob_month": dob_month,
        "dob_year": dob_year,
        "confirmed": False
    }
    
    # Store for real-time updates
    LAST_REGISTRATION = {
        "uid": generated_uid,
        "name": full_name,
        "sex": gender,
        "schedule": class_name,
        "message": f"📝 Awaiting card tap for {full_name}",
        "timestamp": datetime.now().strftime('%H:%M')
    }
    
    print(f"✅ Registration session created - UID: {generated_uid}")
    print(f"   Session ID: {CURRENT_REGISTRATION_SESSION_ID}")
    if photo_file and photo_file.filename:
        print(f"   Photo: {photo_file.filename}")
    return jsonify({
        "status": "waiting", 
        "message": f"Student ID {generated_uid} generated. Please tap card on reader.",
        "uid": generated_uid,
        "session_id": CURRENT_REGISTRATION_SESSION_ID,
        "loading_gif": "/resource/scanning1.gif"
    }), 200

# ======================== THE ENGINE: CHECK-IN & CAPTURE ========================
student_cooldowns = {}  # Track last scan time for each student to enforce 30m rule 
@app.route('/checkin', methods=['GET'])
def checkin():
    global PENDING_REG_DATA, LAST_SCAN_DATA, LAST_REGISTRATION, DUPLICATE_ERROR_TIMESTAMP, PENDING_CONFIRMATIONS
    uid = request.args.get('uid', '').upper()
    session_id = request.args.get('session_id', '')
    now = datetime.now()
    # Detect browser clients so we can redirect after performing check-in
    browser_request = is_browser_request()
    # Also treat requests that originate from the checkin UI (via Referer) as browser requests
    try:
        referer = (request.headers.get('Referer') or request.headers.get('Referrer') or '')
        if '/checkin_page' in referer:
            browser_request = True
    except Exception:
        pass

    def maybe_browser_ok(result_text, default_status=200):
        return (result_text, 200) if browser_request else (result_text, default_status)
    
    if not uid:
        print("⚠️ Warning: Received request with no UID")
        LAST_SCAN_DATA = {
            "uid": None,
            "name": None,
            "status": "UID_MISSING",
            "message": "❌ UID missing from request",
            "sound": "error",
            "timestamp": now.strftime('%H:%M')
        }
        return maybe_browser_ok("UID_MISSING", 400)

    print(f"📡 Card Scanned: {uid}")

    # ===== REGISTRATION CONFIRMATION LOGIC (Single Tap) =====
    # Check if this card tap is for confirming a pending registration
    # If session_id is not provided, auto-detect from PENDING_CONFIRMATIONS
    if not session_id and PENDING_CONFIRMATIONS:
        # Find the first pending registration (should usually be only one)
        session_id = list(PENDING_CONFIRMATIONS.keys())[0]
        print(f"🔍 Auto-detected session_id: {session_id}")
    
    if session_id and session_id in PENDING_CONFIRMATIONS:
        confirmation = PENDING_CONFIRMATIONS[session_id]
        
        # Single tap registration - proceed immediately
        card_uid = uid
        student_id = confirmation["student_id"]
        student_name = confirmation["name"]
        student_sex = confirmation.get('sex', '')
        new_schedule = confirmation.get('schedule', '')
        desk_id = confirmation.get('desk_id', '')
        
        print(f"📝 Single-tap registration: {student_name}")
        print(f"   Card UID: {card_uid}")
        print(f"   Student ID: {student_id}")
        print(f"   Schedule: {new_schedule}")
        
        # Load all students
        students = load_students()
        existing_student = None
        existing_student_idx = None
        
        # Check if this card already exists
        for idx, student in enumerate(students):
            if card_uid in student.get('card_uids', []):
                existing_student = student
                existing_student_idx = idx
                break
        
        if existing_student:
            print(f"🔍 Card already registered to: {existing_student['student_id']} ({existing_student['name']})")
            
            # Check if it's the SAME student (same name + sex) registering for a new class
            if (existing_student['name'].strip() == student_name.strip() and 
                existing_student.get('sex', '') == student_sex):
                
                print(f"✅ SAME STUDENT - Adding new schedule: {new_schedule}")
                
                # Migrate old single-schedule to multi-schedule if needed
                if 'schedule' in existing_student and 'schedules' not in existing_student:
                    existing_student['schedules'] = [existing_student['schedule']]
                    del existing_student['schedule']
                elif 'schedules' not in existing_student:
                    existing_student['schedules'] = []
                
                # Check if already enrolled in this schedule
                if new_schedule not in existing_student['schedules']:
                    existing_student['schedules'].append(new_schedule)
                    
                    # For Computer Class, store desk_id
                    if schedule_is_computer_class(new_schedule) and desk_id:
                        if 'desk_ids' not in existing_student:
                            existing_student['desk_ids'] = {}
                        existing_student['desk_ids'][new_schedule] = desk_id
                    
                    # Update student record
                    students[existing_student_idx] = existing_student
                    save_students(students)
                    
                    print(f"✅ Schedule added successfully. Total schedules: {existing_student['schedules']}")
                else:
                    print(f"⚠️ Student already enrolled in {new_schedule}")
                
                # Clean up and return success
                del PENDING_CONFIRMATIONS[session_id]
                PENDING_REG_DATA = None
                
                LAST_SCAN_DATA = {
                    "uid": existing_student['student_id'],
                    "name": existing_student['name'],
                    "status": "REGISTERED",
                    "message": f"✅ Added '{new_schedule}' to student record",
                    "sound": "success",
                    "timestamp": datetime.now().strftime('%H:%M')
                }
                
                LAST_REGISTRATION = {
                    "student_id": existing_student['student_id'],
                    "name": existing_student['name'],
                    "message": f"✅ Schedule added successfully",
                    "timestamp": datetime.now().strftime('%H:%M')
                }
                return f"SCHEDULE_ADDED_{existing_student['name']}", 200
            
            else:
                # Different student trying to use same card - DUPLICATE CARD ERROR
                print(f"❌ DIFFERENT STUDENT - Duplicate card detected!")
                print(f"   Card belongs to: {existing_student['name']} ({existing_student['student_id']})")
                print(f"   Attempting to register: {student_name} ({student_id})")
                
                DUPLICATE_ERROR_TIMESTAMP = datetime.now()
                LAST_SCAN_DATA = {
                    "uid": student_id,
                    "name": student_name,
                    "status": "DUPLICATE_CARD",
                    "message": f"❌ Card already registered to {existing_student['student_id']}",
                    "sound": "error",
                    "timestamp": datetime.now().strftime('%H:%M'),
                    "error_time": DUPLICATE_ERROR_TIMESTAMP.timestamp()
                }
                del PENDING_CONFIRMATIONS[session_id]
                PENDING_REG_DATA = None
                return "DUPLICATE_CARD_ERROR", 409
        
        else:
            # Brand new card - create new student record
            print(f"✨ New student registration")
            
            new_student = {
                "student_id": student_id,
                "name": student_name,
                "sex": student_sex,
                "schedules": [new_schedule],  # Use schedules array from start
                "card_uids": [card_uid]
            }

            # Record registration date for accurate absence reporting
            try:
                new_student['registered_at'] = datetime.now().strftime('%Y-%m-%d')
            except Exception:
                pass

            # Attach DOB if provided in the pending confirmation (store ISO and parts)
            try:
                dob_day = confirmation.get('dob_day')
                dob_month = confirmation.get('dob_month')
                dob_year = confirmation.get('dob_year')
                if dob_day and dob_month and dob_year:
                    dd = int(dob_day)
                    mm = int(dob_month)
                    yy = int(dob_year)
                    new_student['dob'] = f"{yy:04d}-{mm:02d}-{dd:02d}"
                    new_student['dob_day'] = dd
                    new_student['dob_month'] = mm
                    new_student['dob_year'] = yy
            except Exception as e:
                print(f"⚠️ Could not parse DOB from confirmation: {e}")
            
            # For Computer Class, add desk_id
            if schedule_is_computer_class(new_schedule) and desk_id:
                new_student['desk_ids'] = {new_schedule: desk_id}
            
            # Attach photo if it was provided at registration (already saved)
            photo_filename = confirmation.get('photo')
            if photo_filename:
                new_student['photo'] = photo_filename
                print(f"📷 Using pre-saved photo for student: {photo_filename}")

            # If no photo was uploaded, set default based on sex
            if 'photo' not in new_student or not new_student.get('photo'):
                new_student['photo'] = get_default_photo(student_sex)
            
            students.append(new_student)
            save_students(students)
            
            print(f"✅ New student created with schedule: {new_schedule}")
            
            # Clean up registration data
            del PENDING_CONFIRMATIONS[session_id]
            PENDING_REG_DATA = None
            
            LAST_SCAN_DATA = {
                "uid": student_id,
                "name": student_name,
                "status": "REGISTERED",
                "message": "✅ Student registered successfully",
                "sound": "success",
                "timestamp": datetime.now().strftime('%H:%M'),
                "sex": student_sex,
                "schedule": new_schedule
            }
            
            # Update registration stream
            LAST_REGISTRATION = {
                "student_id": student_id,
                "name": student_name,
                "sex": student_sex,
                "schedule": new_schedule,
                "message": "✅ Student registered successfully",
                "timestamp": datetime.now().strftime('%H:%M')
            }
            return f"REGISTERED_{student_name}", 200
    
    # ===== OLD PENDING_REG_DATA LOGIC (For backwards compatibility, if any) =====
    if PENDING_REG_DATA:
        print(f"⚠️ Warning: Old PENDING_REG_DATA detected, clearing it")
        PENDING_REG_DATA = None

    # 2. FIND STUDENT (Search by CardUID)
    student = find_student_by_card_uid(uid)
    # Note: do NOT redirect here so browser requests still record attendance.
    # Redirect is handled after the attendance record is written below.
        
    if not student:
        print(f"🚫 Unknown Card: {uid}")        
        LAST_SCAN_DATA = {
            "uid": uid,
            "name": "Unknown",
            "status": "UNKNOWN_CARD",
            "message": "❓ Card not registered",
            "sound": "unknown",
            "timestamp": datetime.now().strftime('%H:%M')
        }
        return maybe_browser_ok("UNKNOWN_CARD hahaS", 404)

    # 3. ATTENDANCE LOGIC with TIME WINDOW VALIDATION
    today = datetime.now().strftime('%Y-%m-%d')
    now = datetime.now()
    now_str = now.strftime('%H:%M')
    
    student_id = student['student_id']
    student_name = student['name']
    print(f"✅ Found Student: {student_name} ({student_id})")

    # Enforce short cooldown to prevent rapid check-in/out toggles (2 minutes)
    try:
        cooldown_seconds = 10  # 2 minutes
        last_scan = student_cooldowns.get(student_id)
        if last_scan:
            # last_scan stored as datetime
            seconds_since = (now - last_scan).total_seconds()
            if seconds_since < cooldown_seconds:
                wait = int(cooldown_seconds - seconds_since)
                LAST_SCAN_DATA = {
                    "uid": student_id,
                    "name": student_name,
                    "status": "COOLDOWN",
                    "message": f"⏳ Please wait {wait}s before next scan",
                    "sound": "error",
                    "timestamp": now.strftime('%H:%M')
                }
                return maybe_browser_ok("COOLDOWN", 429)
    except Exception:
        # If cooldown check fails for any reason, continue normally
        pass

    # Get student's primary schedule
    student_schedule = get_primary_schedule(student)
    if not student_schedule:
        print(f"⚠️ Student has no assigned schedule")
        LAST_SCAN_DATA = {
            "uid": student_id,
            "name": student_name,
            "status": "NO_SCHEDULE",
            "message": "⚠️ No schedule assigned",
            "sound": "error",
            "timestamp": now_str
        }
        return maybe_browser_ok("NO_SCHEDULE", 400)

    # Load or create today's record
    attendance_record = None
    row_index = 0
    all_rows = []
    
    try:
        with open(ATTENDANCE_LOG_CSV, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for idx, row in enumerate(reader, start=1):
                all_rows.append(row)
                if row['Date'] == today and row['StudentID'] == student_id:
                    attendance_record = row
                    row_index = idx
    except Exception as e:
        print(f"❌ Attendance Log Read Error: {e}")
        return "SERVER_ERROR", 500

    # Create new record if first scan of the day
    if not attendance_record:
        attendance_record = {
            'Date': today,
            'StudentID': student_id,
            'Name': student_name,
            'In1': '',
            'Out1': '',
            'In2': '',
            'Out2': '',
            'Status': 'IN_SESSION'
        }
        all_rows.append(attendance_record)
        row_index = len(all_rows)

    # ===== DETERMINE NEXT SESSION SLOT & CHECK 5H FLED LIMIT =====
    next_slot = None
    session_num = None

    # Prefer the session that matches current time (so a missed session_1 doesn't block session_2)
    current_session, _ = get_current_session_for_checkin(student, now)

    def _check_5h_and_set_in2():
        nonlocal next_slot, session_num
        # Helper to evaluate 5h limit when moving to In2
        if attendance_record.get('In1'):
            in1_dt = datetime.strptime(f"{today} {attendance_record['In1']}", '%Y-%m-%d %H:%M')
            if (now - in1_dt).total_seconds() / 3600 > 5:
                attendance_record['Status'] = 'FLED'
                print(f"🚩 Session 1 marked FLED (5h limit exceeded)")
                next_slot = 'In2'
                session_num = 2
                return True
        return False

    if current_session == 1:
        # Handle session 1 preference
        if not attendance_record['In1']:
            next_slot = 'In1'
            session_num = 1
        elif attendance_record['In1'] and not attendance_record['Out1']:
            # Check 5-hour limit
            in1_dt = datetime.strptime(f"{today} {attendance_record['In1']}", '%Y-%m-%d %H:%M')
            if (now - in1_dt).total_seconds() / 3600 > 5:
                attendance_record['Status'] = 'FLED'
                print(f"🚩 Session 1 marked FLED (5h limit exceeded)")
                next_slot = 'In2'
                session_num = 2
            else:
                next_slot = 'Out1'
                session_num = 1
        elif attendance_record['In1'] and attendance_record['Out1'] and not attendance_record['In2']:
            next_slot = 'In2'
            session_num = 2
        elif attendance_record['In2'] and not attendance_record['Out2']:
            # Check 5-hour limit for session 2
            in2_dt = datetime.strptime(f"{today} {attendance_record['In2']}", '%Y-%m-%d %H:%M')
            if (now - in2_dt).total_seconds() / 3600 > 5:
                attendance_record['Status'] = 'FLED'
                print(f"🚩 Session 2 marked FLED (5h limit exceeded)")
                LAST_SCAN_DATA = {
                    "uid": student_id,
                    "name": student_name,
                    "status": "FLED",
                    "message": "❌ Session marked FLED: 5h limit exceeded",
                    "sound": "error",
                    "timestamp": now_str,
                    "slot": next_slot
                }
                return maybe_browser_ok("FLED_ERR", 400)
            else:
                next_slot = 'Out2'
                session_num = 2
        elif attendance_record['In2'] and attendance_record['Out2']:
            # Both sessions complete
            print(f"✅ All sessions complete for today")
            LAST_SCAN_DATA = {
                "uid": uid,
                "name": student.get('name', student.get('Name', 'Unknown')),
                "status": "PRESENT",
                "message": "✅ All sessions complete",
                "sound": "success",
                "timestamp": now_str
            }
            return "PRESENT", 200
        else:
            print(f"⚠️ Unexpected slot state")
            next_slot = None
    elif current_session == 2:
        # Handle session 2 when current time is in session 2 window.
        # Allow In2 even if In1 was missed earlier.
        if not attendance_record['In2']:
            next_slot = 'In2'
            session_num = 2
        elif attendance_record['In2'] and not attendance_record['Out2']:
            # Check 5-hour limit for session 2
            in2_dt = datetime.strptime(f"{today} {attendance_record['In2']}", '%Y-%m-%d %H:%M')
            if (now - in2_dt).total_seconds() / 3600 > 5:
                attendance_record['Status'] = 'FLED'
                print(f"🚩 Session 2 marked FLED (5h limit exceeded)")
                LAST_SCAN_DATA = {
                    "uid": student_id,
                    "name": student_name,
                    "status": "FLED",
                    "message": "❌ Session marked FLED: 5h limit exceeded",
                    "sound": "error",
                    "timestamp": now_str,
                    "slot": next_slot
                }
                return maybe_browser_ok("FLED_ERR", 400)
            else:
                next_slot = 'Out2'
                session_num = 2
        elif attendance_record['In2'] and attendance_record['Out2']:
            # If session 2 already complete, fall back to session1 state handling
            pass
        else:
            # If session2 is active but unexpected state, allow In2 as recovery
            next_slot = 'In2'
            session_num = 2

    # Fallback: if we couldn't decide based on current session, keep original sequential behavior
    if not next_slot and session_num is None:
        if not attendance_record['In1']:
            next_slot = 'In1'
            session_num = 1
        elif attendance_record['In1'] and not attendance_record['Out1']:
            # Check 5-hour limit
            in1_dt = datetime.strptime(f"{today} {attendance_record['In1']}", '%Y-%m-%d %H:%M')
            if (now - in1_dt).total_seconds() / 3600 > 5:
                attendance_record['Status'] = 'FLED'
                print(f"🚩 Session 1 marked FLED (5h limit exceeded)")
                next_slot = 'In2'
                session_num = 2
            else:
                next_slot = 'Out1'
                session_num = 1
        elif attendance_record['In1'] and attendance_record['Out1'] and not attendance_record['In2']:
            next_slot = 'In2'
            session_num = 2
        elif attendance_record['In2'] and not attendance_record['Out2']:
            # Check 5-hour limit for session 2
            in2_dt = datetime.strptime(f"{today} {attendance_record['In2']}", '%Y-%m-%d %H:%M')
            if (now - in2_dt).total_seconds() / 3600 > 5:
                attendance_record['Status'] = 'FLED'
                print(f"🚩 Session 2 marked FLED (5h limit exceeded)")
                LAST_SCAN_DATA = {
                    "uid": student_id,
                    "name": student_name,
                    "status": "FLED",
                    "message": "❌ Session marked FLED: 5h limit exceeded",
                    "timestamp": now_str,
                    "slot": next_slot
                }
                return maybe_browser_ok("FLED_ERR", 400)
            else:
                next_slot = 'Out2'
                session_num = 2
        elif attendance_record['In2'] and attendance_record['Out2']:
            # Both sessions complete
            print(f"✅ All sessions complete for today")
            LAST_SCAN_DATA = {
                "uid": uid,
                "name": student.get('name', student.get('Name', 'Unknown')),
                "status": "PRESENT",
                "message": "✅ All sessions complete",
                "sound": "success",
                "timestamp": now_str
            }
            return "PRESENT", 200
        else:
            print(f"⚠️ Unexpected slot state")
            next_slot = None

    # ===== TIME WINDOW VALIDATION =====
    # Allow a browser-based `force=true` override or an admin token to bypass time windows
    token = extract_admin_token()
    is_admin_user = bool(get_admin_user(token))
    force_param = str(request.args.get('force', '')).lower() in ('1', 'true', 'yes')
    force_override = force_param or (is_admin_user and browser_request)

    if next_slot and session_num:
        if "In" in next_slot:
            # CHECKIN TIME WINDOW VALIDATION
            window_result = validate_checkin_window(student_schedule, session_num, now)

            if not window_result['valid'] and not force_override:
                print(f"⚠️ Check-in outside allowed window: {window_result['status']}")
                LAST_SCAN_DATA = {
                    "uid": student_id,
                    "name": student_name,
                    "status": window_result['status'],
                    "message": window_result.get('message', '⏱️ Outside check-in window'),
                    "sound": "error",
                    "timestamp": now_str
                }
                # If this request came from a browser, redirect to the student's report
                if browser_request and student:
                    try:
                        return redirect(url_for('student_report', student_id=student_id, status=window_result['status'], message=window_result.get('message', '')))
                    except Exception:
                        pass
                return maybe_browser_ok(f"OUTSIDE_WINDOW_{window_result['status']}", 400)
            else:
                if force_override and not window_result.get('valid'):
                    print(f"⚠️ Forcing check-in despite window: override active")
                    attendance_status = {'status': 'FORCED', 'message': '✅ Forced check-in (override)'}
                else:
                    print(f"✅ Check-in within allowed window")
                    # Get attendance status (EARLY, ON_TIME, LATE)
                    attendance_status = get_attendance_status(student_schedule, session_num, now_str, now)

        else:
            # CHECKOUT TIME WINDOW VALIDATION
            window_result = validate_checkout_window(student_schedule, session_num, now)

            if not window_result['valid'] and window_result['status'] == 'TOO_EARLY_CHECKOUT' and not force_override:
                # Too early to checkout
                print(f"⚠️ Checkout too early: {window_result['message']}")
                LAST_SCAN_DATA = {
                    "uid": student_id,
                    "name": student_name,
                    "status": "TOO_EARLY_CHECKOUT",
                    "message": window_result['message'],
                    "sound": "error",
                    "timestamp": now_str
                }
                if browser_request and student:
                    try:
                        return redirect(url_for('student_report', student_id=student_id, status='TOO_EARLY_CHECKOUT', message=window_result.get('message', '')))
                    except Exception:
                        pass
                return maybe_browser_ok("TOO_EARLY_CHECKOUT", 429)
            else:
                if force_override and not window_result.get('valid'):
                    print(f"⚠️ Forcing checkout despite window: override active")
                    attendance_status = {'status': 'FORCED', 'message': '✅ Forced checkout (override)'}
                else:
                    print(f"✅ Checkout time validation passed")
                    attendance_status = {'status': window_result['status'], 'message': window_result['message']}
    else:
        print(f"❌ Could not determine slot and session")
        LAST_SCAN_DATA = {
            "uid": student_id if 'student_id' in locals() else uid,
            "name": student_name if 'student_name' in locals() else "Unknown",
            "status": "SLOT_ERROR",
                "message": "❌ Could not determine attendance slot",
                "sound": "error",
            "timestamp": now.strftime('%H:%M')
        }
        return maybe_browser_ok("SLOT_ERROR", 400)

    # NOTE: Removed fixed minimum checkout delay. Checkout allowed per configured time windows.
    # ===== RECORD THE SCAN =====
    if next_slot:
        attendance_record[next_slot] = now_str
        
        # Update status
        if attendance_record['In1'] and attendance_record['Out1'] and attendance_record['In2'] and attendance_record['Out2']:
            attendance_record['Status'] = 'PRESENT'
        else:
            attendance_record['Status'] = 'IN_SESSION'
        
        print(f"📝 Recorded {next_slot}: {now_str}")
    else:
        print(f"❌ Could not determine slot")
        return maybe_browser_ok("SLOT_ERROR", 400)

    # ===== WRITE UPDATED RECORD BACK =====
    try:
        with open(ATTENDANCE_LOG_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['Date', 'StudentID', 'Name', 'In1', 'Out1', 'In2', 'Out2', 'Status'])
            writer.writeheader()
            writer.writerows(all_rows)
    except Exception as e:
        print(f"❌ Attendance Log Write Error: {e}")
        return "SERVER_ERROR", 500

    # Update cooldown timestamp for this student to prevent immediate re-scans
    try:
        student_cooldowns[student_id] = datetime.now()
    except Exception:
        pass

    # ===== UPDATE KIOSK =====
    if "In" in next_slot:
        action_msg = f"Check-in ({attendance_status['status']})"
        detail_msg = attendance_status['message']
    else:
        action_msg = f"Check-out ({attendance_status['status']})"
        detail_msg = attendance_status['message']
    
    if attendance_record['Status'] == 'FLED':
        action_msg = "❌ Session Fled"
        detail_msg = "⏱️ 5h limit exceeded"
    
    # Choose sound: error for FLED or other errors, success otherwise
    sound = 'error' if attendance_record.get('Status') == 'FLED' else 'success'

    LAST_SCAN_DATA = {
        "uid": student_id,
        "name": student_name,
        "status": attendance_record['Status'],
        "message": detail_msg,
        "timestamp": now_str,
        "slot": next_slot,
        "attendance_status": attendance_status.get('status', 'UNKNOWN')
    }
    LAST_SCAN_DATA['sound'] = sound

    # If request came from a browser (or explicit redirect param), redirect to the student's report page.
    # Avoid redirecting ESP32 devices by checking the last known ESP32 IP.
    try:
        client_ip = request.remote_addr
        esp32_ip = PN532_STATUS.get('esp32_ip')
        explicit_redirect = str(request.args.get('redirect', '')).lower() in ('1', 'true', 'yes')
        should_redirect = (browser_request or explicit_redirect) and (client_ip != esp32_ip)
        if should_redirect and student:
            return redirect(url_for('student_report', student_id=student_id))
    except Exception:
        pass

    return "SUCCESS", 200


# Device-friendly scan endpoint: runs the same internal logic as `/checkin`
# but always returns JSON 200 for devices (so ESP32 firmware can treat
# any response as successful while still receiving the status via JSON).
@app.route('/api/scan', methods=['GET'])
def api_scan():
    """Device endpoint: pass `uid` as query param. Internally calls
    `checkin()` to perform the full processing, then returns the
    `LAST_SCAN_DATA` as JSON with HTTP 200 so devices always get 200.
    """
    try:
        # Execute the same processing as /checkin. We ignore the return
        # value because `/checkin` may return redirects or non-200 codes
        # for browsers; the side-effects (attendance write and
        # LAST_SCAN_DATA) are what the device needs.
        _ = checkin()
    except Exception as e:
        print(f"❌ /api/scan internal checkin() error: {e}")

    # Always return 200 with the structured last-scan info
    return jsonify({'ok': True, 'last_scan': LAST_SCAN_DATA}), 200
# ======================== 7-CLASS EXCEL REPORT ========================

@app.route('/report', methods=['GET'])
def report():
    try:
        # Get month/year parameters from query string
        year = request.args.get('year')
        month = request.args.get('month')
        
        wb = Workbook()
        wb.remove(wb.active) # Remove default
        
        # 1. Get Student-to-Schedule Mapping from JSON (includes levels)
        mapping = {}
        students = load_students()
        for student in students:
            # Handle both old format (schedule) and new format (schedules array)
            if 'schedule' in student:
                mapping[student['student_id']] = student['schedule']
            elif 'schedules' in student and student['schedules']:
                # Use first schedule if multiple exist
                mapping[student['student_id']] = student['schedules'][0]
        
        # 2. Collect all unique class/level combinations from students
        unique_schedules = set()
        for schedule in mapping.values():
            if schedule:
                unique_schedules.add(schedule)
        
        # Add default sheets even if no students yet
        default_schedules = [
            "Teacher",
            "Grade 1 - វេនទី 1",
            "Grade 1 - វេនទី 2", 
            "Grade 2 - វេនទី 1",
            "Grade 2 - វេនទី 2",
            "Grade 3 - វេនទី 1",
            "Grade 3 - វេនទី 2",
            "Kindergarten",
            "Computer Class - 17-18pm",
            "Computer Class - 18-19pm",
            "English Class - Essential",
            "English Class - Beginner"
        ]
        
        # Merge student schedules with defaults
        for schedule in default_schedules:
            unique_schedules.add(schedule)
        
        # 3. Create sheets for each unique schedule
        sheets = {}
        for schedule in sorted(unique_schedules):
            # Truncate sheet name if it exceeds Excel's 31 character limit
            sheet_title = schedule if len(schedule) <= 31 else schedule[:28] + "..."
            ws = wb.create_sheet(title=sheet_title)
            ws.append(['Date', 'StudentID', 'Name', 'In 1', 'Out 1', 'In 2', 'Out 2', 'Status'])
            sheets[schedule] = ws

        # 4. Fill Sheets with optional date filtering
        if os.path.exists(ATTENDANCE_LOG_CSV):
            with open(ATTENDANCE_LOG_CSV, 'r', encoding='utf-8') as f:
                for log in csv.DictReader(f):
                    # Filter by month/year if provided
                    if year and month:
                        log_date = log.get('Date', '')
                        if not log_date.startswith(f"{year}-{month.zfill(2)}"):
                            continue
                    
                    schedule = mapping.get(log.get('StudentID'))
                    if schedule and schedule in sheets:
                        sheets[schedule].append([
                            log.get('Date', ''),
                            log.get('StudentID', ''),
                            log.get('Name', ''),
                            log.get('In1', ''),
                            log.get('Out1', ''),
                            log.get('In2', ''),
                            log.get('Out2', ''),
                            log.get('Status', '')
                        ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with month if specified
        filename = "Attendance_Report.xlsx"
        if year and month:
            filename = f"Attendance_Report_{year}-{month.zfill(2)}.xlsx"
        
        return send_file(output, mimetype='application/vnd.ms-excel', as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"❌ Error generating report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# The `serve_resource_file` implementation above handles serving files
# from the `resource/` directory with safety checks and proper mimetypes.

@app.route('/photos/<photo_filename>')
def serve_photo(photo_filename):
    """Serve student photos from the photos directory"""
    # Validate filename to prevent directory traversal attacks
    safe_filename = secure_filename(photo_filename)
    if safe_filename != photo_filename:
        return "Invalid filename", 400
    # First look in the user `photo/` folder
    photo_path = os.path.join(os.path.dirname(__file__), PHOTOS_DIR, safe_filename)
    if os.path.exists(photo_path):
        ext = safe_filename.rsplit('.', 1)[-1].lower() if '.' in safe_filename else ''
        mimetype_map = {
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'webp': 'image/webp',
            'avif': 'image/avif'
        }
        mimetype = mimetype_map.get(ext, 'application/octet-stream')
        return send_file(photo_path, mimetype=mimetype)

    # Fallback: serve default images from the resource/ folder
    resource_path = os.path.join(os.path.dirname(__file__), 'resource', safe_filename)
    if os.path.exists(resource_path):
        ext = safe_filename.rsplit('.', 1)[-1].lower() if '.' in safe_filename else ''
        mimetype_map = {
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'webp': 'image/webp',
            'avif': 'image/avif'
        }
        mimetype = mimetype_map.get(ext, 'application/octet-stream')
        return send_file(resource_path, mimetype=mimetype)

    return "Photo not found", 404

# ======================== REGISTRATION WRITE QUEUE ========================

@app.route('/write/pending', methods=['GET'])
def get_pending_write():
    """
    Get the pending UID that needs to be written to a card.
    Used by both the register page (to know what to wait for) and the simulator (to simulate card writes).
    
    Returns:
        - status: 'pending' if there's a write waiting, 'idle' if not
        - uid: the UID to write (only if status is 'pending')
        - name: student name (only if status is 'pending')
        - session_id: the session ID for this registration attempt (for /checkin to use)
    """
    if PENDING_REG_DATA:
        return jsonify({
            'status': 'pending',
            'uid': PENDING_REG_DATA['uid'],
            'name': PENDING_REG_DATA['name'],
            'sex': PENDING_REG_DATA['sex'],
            'schedule': PENDING_REG_DATA['schedule'],
            'session_id': PENDING_REG_DATA.get('session_id', None),
            'loading_gif': '/resource/scanning1.gif'
        }), 200
    else:
        return jsonify({
            'status': 'idle',
            'uid': None,
            'name': None
        }), 200

@app.route('/send_write_command', methods=['POST'])
def send_write_command():
    """
    Endpoint called by register page to tell ESP32 to enter write mode.
    This is a no-op for the server - it just confirms the command was received.
    The ESP32 will detect pending write via /write/pending polling.
    """
    if PENDING_REG_DATA:
        uid = PENDING_REG_DATA['uid']
        print(f"✍️  Write Mode Activated for UID: {uid}")
        return jsonify({
            'status': 'success',
            'message': f'Write mode activated for {uid}',
            'uid': uid
        }), 200
    else:
        return jsonify({
            'status': 'error',
            'message': 'No pending registration'
        }), 400

@app.route('/write/ack', methods=['POST'])
def write_acknowledge():
    """
    ESP32 calls this endpoint after successfully writing the UID to a card.
    Clears PENDING_REG_DATA to signal completion.
    """
    global PENDING_REG_DATA
    
    data = request.get_json()
    uid_written = data.get('uid') if data else None
    
    if PENDING_REG_DATA and uid_written == PENDING_REG_DATA['uid']:
        print(f"✅ Write Acknowledged: {uid_written} written to card")
        student_data = PENDING_REG_DATA.copy()
        PENDING_REG_DATA = None  # Clear pending write
        
        return jsonify({
            'status': 'success',
            'message': 'Write acknowledged',
            'student': student_data
        }), 200
    else:
        return jsonify({
            'status': 'error',
            'message': 'UID mismatch or no pending write'
        }), 400


@app.route('/api/admin/override', methods=['POST'])
def admin_override():
    """Admin endpoint to manually set attendance slots or status.

    JSON body:
      - admin_token: str
      - student_id: str
      - date: YYYY-MM-DD (optional, default today)
      - slot: In1|Out1|In2|Out2|Status
      - value: string value to set (time for In/Out, or status)
      - reason: optional reason string
    """
    data = request.get_json() or {}
    token = extract_admin_token()
    admin_user = get_admin_user(token)
    if not admin_user:
        return jsonify({'error': 'unauthorized'}), 403
    allowed, rem = check_admin_rate(token)
    if not allowed:
        return jsonify({'error': 'rate_limited'}), 429

    student_id = data.get('student_id')
    if not student_id:
        return jsonify({'error': 'student_id required'}), 400

    date_str = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    slot = data.get('slot')
    value = data.get('value', '')
    reason = data.get('reason', '')

    if slot not in ('In1', 'Out1', 'In2', 'Out2', 'Status'):
        return jsonify({'error': 'invalid slot'}), 400

    # Read all rows
    all_rows = []
    found = False
    try:
        if os.path.exists(ATTENDANCE_LOG_CSV):
            with open(ATTENDANCE_LOG_CSV, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    all_rows.append(row)
        # Find row for date+student
        for row in all_rows:
            if row.get('Date') == date_str and row.get('StudentID') == student_id:
                old_value = row.get(slot, '')
                row[slot] = value
                # Update Status if requested
                if slot == 'Status':
                    row['Status'] = value
                append_audit_entry(admin_user, student_id, date_str, slot, old_value, value, reason)
                found = True
                break

        if not found:
            # Create new row for that date
            new_row = {'Date': date_str, 'StudentID': student_id, 'Name': '', 'In1': '', 'Out1': '', 'In2': '', 'Out2': '', 'Status': 'IN_SESSION'}
            new_row[slot] = value
            if slot == 'Status':
                new_row['Status'] = value
            all_rows.append(new_row)
            append_audit_entry(admin_user, student_id, date_str, slot, '', value, reason)

        # Write back
        with open(ATTENDANCE_LOG_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['Date', 'StudentID', 'Name', 'In1', 'Out1', 'In2', 'Out2', 'Status'])
            writer.writeheader()
            writer.writerows(all_rows)

        # Update kiosk last scan to reflect manual edit
        LAST_SCAN_DATA.update({
            'uid': student_id,
            'name': '',
            'status': 'ADMIN_EDIT',
            'message': f'Manual edit by admin: {slot}={value}',
            'timestamp': datetime.now().strftime('%H:%M'),
            'slot': slot
        })

        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        print(f"❌ Admin override error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/bulk_status', methods=['POST'])
def admin_bulk_status():
    """Apply a status (e.g., EXCUSED, UNUSUAL_EVENT) to multiple students for a date.

    JSON body:
      - admin_token: str
      - status: str (e.g., EXCUSED, UNUSUAL_EVENT)
      - date: YYYY-MM-DD (optional, default today)
      - schedule: str (optional) — class/schedule to apply to; if omitted applies to all students
      - student_ids: [str] (optional) — explicit list of StudentIDs to update
      - reason: optional reason
    """
    data = request.get_json() or {}
    token = extract_admin_token()
    admin_user = get_admin_user(token)
    if not admin_user:
        return jsonify({'error': 'unauthorized'}), 403
    allowed, rem = check_admin_rate(token)
    if not allowed:
        return jsonify({'error': 'rate_limited'}), 429

    status = data.get('status')
    if not status:
        return jsonify({'error': 'status required'}), 400

    date_str = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    schedule = data.get('schedule')
    student_ids = data.get('student_ids') or []
    reason = data.get('reason', '')

    # Load students mapping
    students = load_students()
    student_map = {s.get('student_id'): s for s in students}

    # Decide target student IDs
    targets = set()
    if student_ids:
        for sid in student_ids:
            if sid in student_map:
                targets.add(sid)
    elif schedule:
        for s in students:
            scheds = get_student_schedules(s)
            if schedule in scheds:
                targets.add(s.get('student_id'))
    else:
        # No filter => all students
        for s in students:
            targets.add(s.get('student_id'))

    if not targets:
        return jsonify({'updated': 0, 'message': 'no matching students'}), 200

    # Read attendance log
    all_rows = []
    try:
        if os.path.exists(ATTENDANCE_LOG_CSV):
            with open(ATTENDANCE_LOG_CSV, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    all_rows.append(row)

        updated_count = 0

        # Ensure every target has a row for the date
        existing_map = {(r['Date'], r['StudentID']): r for r in all_rows}

        for sid in targets:
            key = (date_str, sid)
            if key in existing_map:
                row = existing_map[key]
                old = row.get('Status', '')
                row['Status'] = status
                append_audit_entry(admin_user, sid, date_str, 'Status', old, status, reason)
                updated_count += 1
            else:
                # Create a minimal row (Name if available)
                name = student_map.get(sid, {}).get('name', '')
                new_row = {'Date': date_str, 'StudentID': sid, 'Name': name, 'In1': '', 'Out1': '', 'In2': '', 'Out2': '', 'Status': status}
                all_rows.append(new_row)
                append_audit_entry(admin_user, sid, date_str, 'Status', '', status, reason)
                updated_count += 1

        # Write back attendance CSV
        with open(ATTENDANCE_LOG_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['Date', 'StudentID', 'Name', 'In1', 'Out1', 'In2', 'Out2', 'Status'])
            writer.writeheader()
            writer.writerows(all_rows)

        # Update kiosk stream with summary
        LAST_SCAN_DATA.update({
            'uid': None,
            'name': '',
            'status': 'ADMIN_BULK',
            'message': f'Bulk set {status} for {updated_count} students',
            'timestamp': datetime.now().strftime('%H:%M')
        })

        return jsonify({'updated': updated_count}), 200
    except Exception as e:
        print(f"❌ Bulk status error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/admin')
def admin_page():
    return render_template('admin.html')


@app.route('/api/admin/keys')
def api_admin_keys():
    token = extract_admin_token()
    admin_user = get_admin_user(token)
    if not admin_user:
        return jsonify({'error': 'unauthorized'}), 403
    allowed, rem = check_admin_rate(token)
    if not allowed:
        return jsonify({'error': 'rate_limited'}), 429

    keys = load_persistent_admin_keys()
    # Return id, username, created_at only
    out = [{'id': k.get('id'), 'username': k.get('username'), 'created_at': k.get('created_at')} for k in keys]
    return jsonify(out)


@app.route('/api/admin/keys/create', methods=['POST'])
def api_admin_keys_create():
    data = request.get_json() or {}
    token = extract_admin_token()
    admin_user = get_admin_user(token)
    if not admin_user:
        return jsonify({'error': 'unauthorized'}), 403
    allowed, rem = check_admin_rate(token)
    if not allowed:
        return jsonify({'error': 'rate_limited'}), 429

    username = data.get('username') or 'admin'
    # Generate token and store hashed
    new_token = secrets.token_urlsafe(24)
    token_hash = hashlib.sha256(new_token.encode()).hexdigest()
    entry = {'id': str(uuid.uuid4()), 'username': username, 'token_hash': token_hash, 'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    keys = load_persistent_admin_keys()
    keys.append(entry)
    save_persistent_admin_keys(keys)

    # Return plaintext token once
    return jsonify({'id': entry['id'], 'username': username, 'token': new_token}), 200


@app.route('/api/admin/keys/revoke', methods=['POST'])
def api_admin_keys_revoke():
    data = request.get_json() or {}
    token = extract_admin_token()
    admin_user = get_admin_user(token)
    if not admin_user:
        return jsonify({'error': 'unauthorized'}), 403
    allowed, rem = check_admin_rate(token)
    if not allowed:
        return jsonify({'error': 'rate_limited'}), 429

    key_id = data.get('key_id')
    if not key_id:
        return jsonify({'error': 'key_id required'}), 400

    keys = load_persistent_admin_keys()
    new_keys = [k for k in keys if k.get('id') != key_id]
    if len(new_keys) == len(keys):
        return jsonify({'error': 'key not found'}), 404
    save_persistent_admin_keys(new_keys)
    return jsonify({'status': 'revoked', 'id': key_id}), 200


@app.route('/api/admin/audit')
def api_admin_audit():
    token = extract_admin_token()
    admin_user = get_admin_user(token)
    if not admin_user:
        return jsonify({'error': 'unauthorized'}), 403
    allowed, rem = check_admin_rate(token)
    if not allowed:
        return jsonify({'error': 'rate_limited'}), 429

    if not os.path.exists(AUDIT_LOG_CSV):
        return jsonify([])

    try:
        with open(AUDIT_LOG_CSV, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = [r for r in reader]
        return jsonify(rows)
    except Exception as e:
        print(f"❌ Audit read error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/audit_csv')
def api_admin_audit_csv():
    token = extract_admin_token()
    admin_user = get_admin_user(token)
    if not admin_user:
        return jsonify({'error': 'unauthorized'}), 403
    allowed, rem = check_admin_rate(token)
    if not allowed:
        return jsonify({'error': 'rate_limited'}), 429

    if not os.path.exists(AUDIT_LOG_CSV):
        return "", 404

    try:
        return send_file(AUDIT_LOG_CSV, mimetype='text/csv', as_attachment=True, download_name='attendance_audit.csv')
    except Exception as e:
        print(f"❌ Audit CSV send error: {e}")
        return jsonify({'error': str(e)}), 500

# ======================== REPLACEMENT CARD REGISTRATION ========================

@app.route('/register_replacement_card', methods=['POST'])
def register_replacement_card():
    """
    Register a new/replacement card for an existing student.
    The student will tap their new card, and this endpoint will add the new CardUID to their record.
    
    Expected flow:
    1. Admin finds student by StudentID
    2. Student taps new card on reader
    3. ESP32 calls this endpoint with the new CardUID
    4. System adds new CardUID to student's card_uids array
    """
    global PENDING_REG_DATA
    
    data = request.get_json()
    new_card_uid = data.get('card_uid', '').upper() if data else ''
    student_id = data.get('student_id', '').upper() if data else ''
    
    if not new_card_uid or not student_id:
        return jsonify({'status': 'error', 'message': 'Missing card_uid or student_id'}), 400
    
    students = load_students()
    student_found = None
    
    for student in students:
        if student['student_id'] == student_id:
            student_found = student
            break
    
    if not student_found:
        return jsonify({'status': 'error', 'message': f'Student {student_id} not found'}), 404
    
    # Check if card already registered to this or another student
    for student in students:
        if new_card_uid in student.get('card_uids', []):
            if student['student_id'] == student_id:
                return jsonify({
                    'status': 'warning',
                    'message': f'Card already registered to {student_id}'
                }), 200
            else:
                return jsonify({
                    'status': 'error',
                    'message': f'Card already registered to {student["student_id"]}'
                }), 409
    
    # Add new card UID to student
    if new_card_uid not in student_found.get('card_uids', []):
        student_found['card_uids'].append(new_card_uid)
        save_students(students)
        
        print(f"✅ New card registered for {student_id}: {new_card_uid}")
        print(f"   Total cards for this student: {len(student_found['card_uids'])}")
        
        return jsonify({
            'status': 'success',
            'message': f'Card registered for {student_id}',
            'student_id': student_id,
            'student_name': student_found['name'],
            'card_count': len(student_found['card_uids'])
        }), 200
    
    return jsonify({'status': 'error', 'message': 'Failed to add card'}), 500


@app.route('/ict')
def ict_desk_map():
    """Render the ICT desk map page."""
    return render_template('ICT_deskMap.html')


if __name__ == '__main__':
    start_mdns()
    app.run(host='0.0.0.0', port=8080)
